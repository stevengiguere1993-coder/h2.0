"""Extraction d'infos immeuble — parser local + Gemini EN PARALLÈLE.

Pipeline (Phase A1, refonte 2026-05-22) — stratégie « sans faille » :
  1. COUCHE 1 — parser Python pur (regex + parsers spécifiques
     Centris/DuProprio/Realtor + JSON-LD + __NEXT_DATA__ + OCR
     Tesseract). Rapide, gratuit, sans quota, hors-ligne.
  2. COUCHE 2 — Gemini 2.0 Flash (palier gratuit) lancé EN PARALLÈLE
     du parser local via ``asyncio.gather`` (et non plus en cascade
     en filet de secours). Reçoit TOUS les inputs (HTML strippé,
     texte libre, images, PDFs OCR-isés) et complète/corrige le
     résultat local par un merge intelligent par champ.
  3. COUCHE 3 — Claude Sonnet (sera ajoutée en Phase A2 sur un
     bouton dédié — pas encore active ici).

Merge intelligent (par champ) :
  - concordance des 2 sources (±5% pour numériques, strings
    normalisées identiques) → on garde la valeur locale (confiance
    forte).
  - parser local seul → on garde le parser local (les parsers
    spécifiques Centris/DuProprio/Realtor sont fiables).
  - Gemini seul → on garde Gemini (le parser local a raté ce champ).
  - divergence majeure → on garde Gemini par défaut + on émet un
    warning visible « Divergence sur {champ} … vérifier
    manuellement ». Exception : address/city/postal_code clairement
    valides côté local (commence par chiffre, ville reconnue) → on
    garde local malgré la divergence.

`model_used` indique la cascade réelle utilisée pour cette
extraction : ``"local + gemini"`` (les deux ont produit), ``"local"``
(Gemini désactivé, en erreur, ou n'a rien renvoyé), ``"gemini"`` (le
parser local n'a rien sorti) ou ``"none"`` (les deux vides).

Sources supportées :
  - URLs Centris.ca         → parser CSS spécifique (BeautifulSoup)
  - URLs DuProprio.com      → parser CSS spécifique
  - URLs Realtor.ca         → parser CSS spécifique
  - URLs Pmml.ca / Next.js  → bloc __NEXT_DATA__ (JSON applicatif)
  - URLs génériques         → JSON-LD (schema.org) en priorité,
                              sinon regex best-effort sur HTML stripped
  - Texte libre             → regex + heuristiques (prix, logements,
                              adresse, code postal, évaluation,
                              taxes, courtier, typologie québécoise)
  - PDFs texte              → pypdf + regex texte
  - PDFs scannés            → fallback OCR Tesseract (pdf2image →
                              images PIL → pytesseract page par page)
  - Images (PNG/JPG/HEIC/…) → OCR Tesseract (pytesseract sur l'image
                              décodée par Pillow / pillow-heif)
  - Excel (.xlsx/.xls)      → openpyxl → texte structuré (headers +
                              lignes/colonnes séparées par « | »)
                              passé à parse_text + Gemini

API publique inchangée : `extract_lead_info(urls, text, files)` →
`ExtractionResult(data, model_used, raw_response, warnings)`.
La structure JSON renvoyée au frontend ne change pas — seul le
contenu de ``model_used`` est enrichi (cf. plus haut) et il y a
potentiellement plus de warnings utiles (divergences, Gemini
indisponible, etc.).

Autres services du repo qui restent sur leur LLM (hors scope de cette
refonte) : `estimate-expenses` (Claude) et `debug-extract-url`
(garde Gemini pour debug d'extraction). Les paquets
`google-generativeai` et `anthropic` restent donc dans les deps.

Stack OCR (binaires système installés via backend/Aptfile sur Render) :
  - tesseract-ocr + tesseract-ocr-fra  → moteur OCR + pack français
  - poppler-utils                      → pdftoppm pour pdf2image
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import re
import time
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import httpx

from app.core.config import settings
from app.integrations import scraping_proxy

log = logging.getLogger(__name__)


# Tag legacy — conservé pour compat éventuelle. Le pipeline retourne
# désormais des valeurs enrichies (``"local + gemini"``, ``"local"``,
# ``"gemini"``, ``"none"``) via _select_model_used().
MODEL_TAG = "local-parser-v1"


# Tolérance utilisée par _values_concordant() : pour les champs
# numériques, on considère deux valeurs concordantes si elles sont à
# moins de 5% l'une de l'autre (couvre les petits écarts de parsing,
# arrondis, conversions M/k, etc.).
_NUMERIC_TOLERANCE = 0.05


# Champs adresse/géographie sur lesquels on préfère la valeur locale
# si elle est « clairement valide » (adresse qui commence par un
# chiffre, code postal canadien) — même en cas de divergence avec
# Gemini, qui hallucine parfois sur l'adresse civique.
_GEO_FIELDS = {"address", "city", "postal_code", "province"}


# Champs traités comme numériques pour la comparaison de concordance.
_NUMERIC_FIELDS = {
    "asking_price", "nb_logements", "revenus_bruts",
    "taxes_municipales", "taxes_scolaires", "assurances", "energie",
    "depenses_autres", "annee_construction", "superficie_terrain",
    "superficie_batiment", "evaluation_municipale",
    "nb_stationnements",
}


# Champs « clés » d'une fiche immeuble — sert à mesurer la couverture
# d'une extraction. Très peu de champs trouvés ⇒ format mal pris en
# charge : on le signale pour qu'il soit transmis et le parser amélioré.
_KEY_FIELDS: Tuple[str, ...] = (
    "address", "city", "asking_price", "nb_logements", "typology",
    "revenus_bruts", "taxes_municipales", "taxes_scolaires",
    "assurances", "energie", "evaluation_municipale",
    "annee_construction", "superficie_terrain", "superficie_batiment",
)


def _coverage(rec: Dict[str, Any]) -> int:
    """Nombre de champs clés réellement renseignés dans une fiche."""
    return sum(
        1 for f in _KEY_FIELDS if rec.get(f) not in (None, "", {}, [])
    )


# ── Activation du support HEIC/HEIF (photos iPhone) ───────────────
#
# Depuis iOS 11, les iPhones prennent des photos en .heic par défaut.
# Pillow standard ne sait pas lire ce format — il faut enregistrer un
# opener via pillow-heif. L'import est paresseux pour ne pas planter
# si le paquet est absent en local (dev), mais en prod (Render) il
# est installé via requirements.txt.
try:
    from pillow_heif import register_heif_opener  # type: ignore

    register_heif_opener()
except ImportError:  # pragma: no cover — env dev sans pillow-heif
    log.info("pillow-heif absent — support HEIC/HEIF désactivé")


# Seuil sous lequel on considère qu'un PDF est "scanné" (couche texte
# vide ou bruitée) et qu'on tombe en fallback OCR. 50 caractères c'est
# moins qu'une ligne d'adresse + prix — n'importe quel PDF descriptif
# réel dépasse largement.
_PDF_OCR_FALLBACK_THRESHOLD = 50


# Normalisations OCR fréquentes. Tesseract confond parfois certains
# caractères sur du texte stylisé. Ces remplacements sont conservateurs
# (ne touchent QUE des contextes ambigus) — on les applique APRÈS l'OCR
# mais AVANT parse_text(), pour donner toutes les chances aux regex.
_OCR_FIXES = [
    # « S 1 234 » → « $ 1 234 » (Tesseract rend souvent $ comme S quand
    # la police est fine). Conditionné à un nombre qui suit.
    (re.compile(r"\bS\s+(\d[\d\s\.,]*)"), r"$ \1"),
    # Caractères pipe verticaux en bordure de tableau → espace.
    (re.compile(r"[|]+"), " "),
    # Ligatures et caractères mal reconnus dans les nombres : « O »
    # entre deux chiffres → « 0 » ; « l » entre deux chiffres → « 1 ».
    (re.compile(r"(\d)\s*[Oo]\s*(\d)"), r"\1 0 \2"),
    (re.compile(r"(\d)\s*[lI]\s*(\d)"), r"\1 1 \2"),
]


def _normalize_ocr_text(text: str) -> str:
    """Applique des corrections best-effort au texte sorti par Tesseract.

    Tesseract peut introduire des confusions sur du texte stylisé
    (chiffres ambigus, dollar mal reconnu). On nettoie avant de passer
    à parse_text() pour donner toutes les chances aux regex. Restera
    toujours du bruit résiduel — c'est attendu, le parser est tolérant.
    """
    if not text:
        return ""
    out = text
    for pattern, repl in _OCR_FIXES:
        out = pattern.sub(repl, out)
    return out


def _check_tesseract_status() -> str:
    """Vérifie si Tesseract est installé et lit la version. Pour
    diagnostic dans les warnings utilisateur quand l'OCR retourne
    du vide."""
    try:
        import pytesseract  # type: ignore
        version = pytesseract.get_tesseract_version()
        return f"OK (Tesseract v{version})"
    except ImportError:
        return "pytesseract non installé"
    except Exception as exc:  # noqa: BLE001
        return f"binaire absent ou erreur : {exc}"


def parse_image_ocr(image_bytes: bytes, filename: str = "image") -> str:
    """OCR sur une image (PNG/JPEG/HEIC/etc.). Retourne le texte extrait.

    Stack : Pillow décode l'image (HEIC géré via pillow-heif registré
    au module load), puis pytesseract appelle le binaire Tesseract en
    français+anglais. Phil reçoit souvent des screenshots de tableaux
    Excel (texte propre, contraste fort) — Tesseract performe très
    bien sur ce type d'input.

    Retourne une chaîne vide si l'OCR échoue (image illisible, binaire
    Tesseract absent, exception inattendue). Le caller émet alors un
    warning explicite.
    """
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore
    except ImportError as exc:
        log.warning("OCR désactivé — paquet manquant : %s", exc)
        return ""

    t0 = time.perf_counter()
    try:
        img = Image.open(io.BytesIO(image_bytes))
        # Tesseract préfère RGB/L. Pour HEIC, RGBA, palette indexée, on
        # convertit pour éviter "OSError: cannot write mode RGBA as JPEG"
        # ou des résultats dégradés.
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        text = pytesseract.image_to_string(img, lang="fra+eng") or ""
    except Exception as exc:  # noqa: BLE001
        log.warning("OCR image '%s' a échoué : %s", filename, exc)
        return ""
    dt = time.perf_counter() - t0
    log.info(
        "OCR image '%s' : %d chars extraits en %.2fs",
        filename,
        len(text),
        dt,
    )
    return text


def parse_pdf_ocr(pdf_bytes: bytes, filename: str = "pdf") -> str:
    """OCR sur un PDF scanné. Convertit en images puis OCR page par page.

    Utilisé en fallback quand pypdf retourne une chaîne vide ou trop
    courte (PDF purement scanné, sans couche texte). Nécessite le
    binaire `pdftoppm` fourni par poppler-utils (installé via Aptfile
    sur Render).

    DPI 200 est un bon compromis qualité/vitesse pour du texte de
    fiche MLS scannée. Au-delà la précision n'augmente plus mais le
    temps explose (1-5 sec par page à 200 DPI, 10+ sec à 400 DPI).
    """
    try:
        import pytesseract  # type: ignore
        from pdf2image import convert_from_bytes  # type: ignore
    except ImportError as exc:
        log.warning("OCR PDF désactivé — paquet manquant : %s", exc)
        return ""

    t0 = time.perf_counter()
    try:
        images = convert_from_bytes(pdf_bytes, dpi=200)
    except Exception as exc:  # noqa: BLE001
        log.warning(
            "Conversion PDF→images (poppler) a échoué pour '%s' : %s",
            filename,
            exc,
        )
        return ""

    texts: List[str] = []
    for i, img in enumerate(images, start=1):
        try:
            page_text = pytesseract.image_to_string(img, lang="fra+eng") or ""
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "OCR PDF '%s' page %d a échoué : %s", filename, i, exc
            )
            continue
        if page_text.strip():
            texts.append(page_text)

    full = "\n\n".join(texts)
    dt = time.perf_counter() - t0
    log.info(
        "OCR PDF '%s' : %d pages, %d chars extraits en %.2fs",
        filename,
        len(images),
        len(full),
        dt,
    )
    return full


# ── Compatibilité : exports utilisés par `debug-extract-url` ──────
#
# L'endpoint diagnostic ``/lead-analyses/debug-extract-url`` reste sur
# Gemini pour permettre à l'utilisateur de comparer ce que voit le
# LLM vs ce que voient les parsers locaux. Pour qu'il continue de
# fonctionner sans modifier l'endpoint, on ré-exporte ici les
# symboles attendus. La refonte du pipeline principal n'a aucun lien
# avec le debug — ils coexistent.

import os as _os

EXTRACTION_MODEL = _os.environ.get(
    "LEAD_EXTRACTION_MODEL", "gemini-2.0-flash"
)

SYSTEM_PROMPT = (
    "Tu es un assistant d'extraction de données pour un dirigeant "
    "qui acquiert des immeubles à logements au Québec. Tu reçois des "
    "sources hétérogènes (Centris, DuProprio, PMML, courriel, photo "
    "de fiche MLS, PDF descriptif, capture d'écran, SMS). Convertis "
    "TOUS les nombres en valeur numérique pure. Pour la typologie, "
    "parse en dict { \"4.5\": 4, \"5.5\": 8 }. Adresse civique dans "
    "`address`, ville dans `city`. Si une info n'est pas présente, "
    "retourne null. Réponds UNIQUEMENT avec du JSON strict."
)

SCHEMA_GUIDE = (
    "Schéma JSON attendu (1 objet ou array si plusieurs immeubles) :\n"
    "{ \"address\": str, \"city\": str, \"postal_code\": str, "
    "\"province\": str, \"asking_price\": int, \"nb_logements\": int, "
    "\"typology\": { \"3.5\": int, \"4.5\": int, ... }, "
    "\"revenus_bruts\": int, \"taxes_municipales\": int, "
    "\"taxes_scolaires\": int, \"assurances\": int, \"energie\": int, "
    "\"depenses_autres\": int, \"annee_construction\": int, "
    "\"superficie_terrain\": int, \"superficie_batiment\": int, "
    "\"evaluation_municipale\": int, \"description\": str, "
    "\"courtier_nom\": str, \"courtier_contact\": str, "
    "\"type_batiment\": str, \"nb_stationnements\": int }"
)


async def _fetch_url_text(url: str) -> str:
    """Variante texte du fetch — utilisée par l'endpoint debug pour
    montrer ce que le LLM verrait. Renvoie le HTML stripé + meta tags
    + JSON-LD + __NEXT_DATA__ concaténés (lisible)."""
    html, err = await _fetch_html(url)
    if err:
        return f"[URL non accessible : {err}] {url}"
    parts: List[str] = [f"[Source URL : {url}]"]
    nd = _NEXTDATA_RE.search(html)
    if nd:
        chunk = nd.group(1).strip()
        if len(chunk) > 40_000:
            chunk = chunk[:40_000] + "\n[…tronqué]"
        parts.append(f"[Bloc __NEXT_DATA__]\n{chunk}")
    for i, ld in enumerate(_JSONLD_RE.findall(html)[:5]):
        clean = ld.strip()
        if len(clean) > 8_000:
            clean = clean[:8_000] + "\n[…tronqué]"
        if clean:
            parts.append(f"[JSON-LD #{i + 1}]\n{clean}")
    stripped = _strip_html(html)
    if len(stripped) > 35_000:
        stripped = stripped[:35_000] + "\n[…tronqué]"
    parts.append(f"[Texte de la page]\n{stripped}")
    return "\n\n".join(parts)


# ── Extraction IA (Gemini) — primaire du pipeline hybride ─────────
#
# Le parser regex local est rapide et autonome mais fragile : il
# confond l'année et le montant des taxes, rate la typologie, etc.
# Gemini *comprend* la fiche et extrait correctement. On l'utilise
# donc en PRIMAIRE, avec le parser local en filet de secours si
# Gemini est indisponible (clé absente, quota atteint, panne réseau).
# L'extraction fonctionne ainsi toujours, même hors-ligne.

_GEMINI_BASE = "https://generativelanguage.googleapis.com/v1beta"


async def _gemini_extract(
    material: str,
    images: List[Tuple[str, bytes]],
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
    """Extrait les champs immeuble via Gemini.

    Retourne ``(data, None)`` en cas de succès, ou ``(None, raison)``
    sinon — où ``raison`` est un message lisible expliquant pourquoi
    Gemini n'a pas été utilisé (affiché à l'utilisateur comme
    avertissement, pour diagnostiquer sans fouiller les logs)."""
    api_key = (getattr(settings, "gemini_api_key", None) or "").strip()
    if not api_key:
        log.warning("Gemini : GEMINI_API_KEY absente — parser local")
        return None, "clé GEMINI_API_KEY absente du serveur"
    if not material.strip() and not images:
        return None, None

    user_parts: List[Dict[str, Any]] = [{"text": SCHEMA_GUIDE}]
    if material.strip():
        user_parts.append(
            {"text": "Sources à analyser :\n\n" + material[:60_000]}
        )
    for mime, blob in images[:4]:  # garde-fou : 4 images max
        if not blob:
            continue
        user_parts.append(
            {
                "inlineData": {
                    "mimeType": mime or "image/png",
                    "data": base64.standard_b64encode(blob).decode("ascii"),
                }
            }
        )

    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": [{"role": "user", "parts": user_parts}],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
            "maxOutputTokens": 4096,
        },
    }
    url = f"{_GEMINI_BASE}/models/{EXTRACTION_MODEL}:generateContent"
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                url, params={"key": api_key}, json=payload
            )
        if resp.status_code == 429:
            log.warning("Gemini : quota atteint — fallback parser local")
            return None, "quota Gemini atteint"
        if resp.status_code >= 400:
            log.warning(
                "Gemini extraction HTTP %s : %s",
                resp.status_code,
                resp.text[:300],
            )
            return None, f"erreur Gemini HTTP {resp.status_code}"
        body = resp.json()
        text = (
            (body.get("candidates") or [{}])[0]
            .get("content", {})
            .get("parts", [{}])[0]
            .get("text", "")
        )
        if not text.strip():
            return None, "réponse Gemini vide"
        parsed = json.loads(text)
    except Exception as exc:  # noqa: BLE001
        log.warning("Gemini extraction échouée : %s", exc)
        return None, f"Gemini injoignable ({type(exc).__name__})"

    items = parsed if isinstance(parsed, list) else [parsed]
    out: List[Dict[str, Any]] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        # Retire les champs vides — `null`, "", "null" — pour ne pas
        # écraser inutilement (le formulaire reste blanc si absent).
        clean = {
            k: v for k, v in it.items() if v not in (None, "", "null")
        }
        if clean:
            out.append(clean)
    if not out:
        return None, "Gemini n'a renvoyé aucun champ"
    return out, None


# ── Dataclasses publiques ─────────────────────────────────────────


@dataclass
class ExtractionInput:
    """Un input à extraire (compat avec l'ancienne signature)."""

    url: Optional[str] = None
    text: Optional[str] = None
    # (filename, content_type, raw bytes)
    file_data: Optional[Tuple[str, str, bytes]] = None


@dataclass
class ExtractionResult:
    """Résultat d'une extraction. `data` est une liste de dicts (un
    par immeuble distinct détecté — souvent un seul).

    ``per_source_values`` (Phase A3, validation post-extraction)
    contient, POUR CHAQUE IMMEUBLE de ``data`` (même ordre, même
    index), un mapping ``{field: {"local": ..., "gemini": ...}}``
    qui conserve la valeur vue par chaque couche AVANT le merge.
    Sert au service ``lead_validation`` pour détecter les divergences
    et alimenter le tooltip côté UI ("Local : X / Gemini : Y").
    Reste vide ``[]`` si une seule couche a contribué.
    """

    data: List[dict] = field(default_factory=list)
    model_used: Optional[str] = None
    raw_response: Optional[str] = None
    warnings: List[str] = field(default_factory=list)
    per_source_values: List[Dict[str, Dict[str, Any]]] = field(
        default_factory=list
    )


# ── Helpers numériques ────────────────────────────────────────────


_NUM_CLEAN_RE = re.compile(r"[^\d.,kKmM]")


def normalize_number(s: Any) -> Optional[float]:
    """Convertit une valeur potentiellement bruitée en float.

    Robuste aux formats vus en immobilier québécois :
      - "1 200 000"      → 1_200_000.0
      - "1,200,000"      → 1_200_000.0
      - "1.2M" / "1,2 M" → 1_200_000.0
      - "1.5k"           → 1_500.0
      - "3 560 000 $"    → 3_560_000.0
      - "84 000"         → 84_000.0
      - 1250000 (int)    → 1_250_000.0

    Retourne ``None`` si la valeur n'est pas convertible.
    """
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    raw = str(s).strip()
    if not raw:
        return None

    # Détecte un suffixe M / k (multiplicateur).
    multiplier = 1.0
    last = raw[-1]
    if last in "Mm":
        multiplier = 1_000_000.0
        raw = raw[:-1].strip()
    elif last in "Kk":
        multiplier = 1_000.0
        raw = raw[:-1].strip()

    # Vire tout ce qui n'est pas chiffre/virgule/point.
    cleaned = re.sub(r"[^\d.,]", "", raw)
    if not cleaned:
        return None

    # Heuristique séparateur décimal :
    # - Si la chaîne contient à la fois "," et "." : le DERNIER des deux
    #   est le séparateur décimal (les autres sont des milliers).
    # - Si elle contient un seul "," et au plus 2 chiffres après, c'est
    #   décimal (style FR : "1,5").
    # - Sinon les "," et "." sont des séparateurs de milliers.
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        parts = cleaned.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    # "." seul : ambigu (peut être milliers FR rares ou décimal EN).
    # On suppose décimal — sauf si 3 chiffres exactement après et pas
    # de M/k (style "1.250" mal formaté → 1250).
    elif "." in cleaned and multiplier == 1.0:
        parts = cleaned.split(".")
        if (
            len(parts) == 2
            and len(parts[1]) == 3
            and len(parts[0]) <= 3
        ):
            # Probable séparateur de milliers européen rare.
            cleaned = cleaned.replace(".", "")

    try:
        return float(cleaned) * multiplier
    except ValueError:
        return None


def _int_or_none(s: Any) -> Optional[int]:
    """Variante de normalize_number qui force le résultat en int."""
    v = normalize_number(s)
    if v is None:
        return None
    try:
        return int(round(v))
    except (ValueError, OverflowError):
        return None


# ── Parser texte libre (regex + heuristiques) ─────────────────────


_POSTAL_RE = re.compile(r"\b([A-Z]\d[A-Z])\s?(\d[A-Z]\d)\b", re.I)
_PHONE_RE = re.compile(
    r"(?:\+?1[\-.\s]?)?\(?(\d{3})\)?[\-.\s]?(\d{3})[\-.\s]?(\d{4})"
)
_EMAIL_RE = re.compile(r"[\w\.\-+]+@[\w\.\-]+\.[a-z]{2,}", re.I)


def parse_text(text: str) -> Dict[str, Any]:
    """Extrait des champs depuis du texte libre (description courtier,
    SMS, copier-coller, texte de PDF). Best-effort, retourne un dict
    partiel — jamais d'exception non catchée."""
    if not text:
        return {}
    # Normalise les apostrophes typographiques — Centris écrit
    # « Nombre d'unités » avec ' (U+2019), pas l'apostrophe droite.
    # Sans ça, les regex en « ' » échouent (ex. nb logements raté).
    text = text.replace("’", "'").replace("‘", "'")
    out: Dict[str, Any] = {}

    # Prix demandé. Plusieurs formes : "Prix demandé : 3 560 000 $",
    # "Prix demandé 1,2M$", "1 250 000 $". On préfère la forme
    # qualifiée (avec "demandé") quand elle existe.
    m = re.search(
        r"(?:prix\s+demand[eé]|prix\s+de\s+vente|asking\s+price)[\s:]*\$?\s*"
        r"([\d][\d\s\.,]*[\d])\s*(?:\$|cad)?",
        text,
        flags=re.I,
    )
    if not m:
        m = re.search(
            r"([\d][\d\s\.,]*\d)\s*(?:M|m)\s*\$",  # "1,2 M$"
            text,
        )
        if m:
            v = normalize_number(m.group(1) + "M")
            if v:
                out["asking_price"] = int(round(v))
        else:
            # Fallback générique : un montant > 100 000 $ proche du
            # mot "prix" → asking_price probable.
            m = re.search(
                r"prix[^\d]{0,40}?([\d][\d\s\.,]{4,})\s*\$",
                text,
                flags=re.I,
            )
            if m:
                v = _int_or_none(m.group(1))
                if v and v >= 50_000:
                    out["asking_price"] = v
    else:
        v = normalize_number(m.group(1))
        if v:
            # Cas "1,2 M $" capté dans le 1er regex aussi.
            tail = text[m.end():m.end() + 6].lower()
            if "m" in tail and v < 1000:
                v = v * 1_000_000
            out["asking_price"] = int(round(v))

    # Nombre de logements. "12 logements", "12 unités", "12 portes",
    # "Nombre de logements : 8".
    m = re.search(
        r"nombre\s+d['e]\s*(?:logements|unit[eé]s)[^\d]{0,15}?(\d{1,3})",
        text,
        flags=re.I,
    )
    if m:
        out["nb_logements"] = int(m.group(1))
    else:
        m = re.search(
            r"\b(\d{1,3})\s*(?:logements?|unit[eé]s?|portes?)\b",
            text,
            flags=re.I,
        )
        if m:
            v = int(m.group(1))
            if 1 <= v <= 999:
                out["nb_logements"] = v

    # Typologie québécoise (X.5 ou X ½, caractéristique 1.5, 2.5, 3.5,
    # 4.5, 5.5, 6.5). Pattern : "8 x 5.5", "12 x 4 ½", "8 unités de
    # 5½". On REFUSE les patterns sans ".5" (faux positifs : "2 x 6").
    typology: Dict[str, int] = {}
    # Accepte « 3 ½ », « 3.5 », « 3,5 » ET « 3 1/2 » (Centris écrit
    # la typologie en ½ dans le tableau et en « 1/2 » dans le texte).
    typo_re = re.compile(
        r"(\d{1,2})\s*(?:[xX×]|unit[eé]s?\s+de)\s*(\d)"
        r"\s*(?:[\.,]\s*5|½|1\s*/\s*2)",
    )
    for m in typo_re.finditer(text):
        try:
            qty = int(m.group(1))
            base = int(m.group(2))
            key = f"{base}.5"
            # On garde le MAX si le texte cite plusieurs fois la même
            # typologie (description + tableau).
            typology[key] = max(typology.get(key, 0), qty)
        except ValueError:
            pass
    if typology:
        out["typology"] = typology
        # Si nb_logements absent, déduit de la typologie.
        if "nb_logements" not in out:
            out["nb_logements"] = sum(typology.values())

    # Année de construction. "Bâti en 1965", "Année de construction 1908",
    # "construit en 1972".
    m = re.search(
        r"(?:b[aâ]ti(?:e)?(?:\s+en)?|construit(?:e)?(?:\s+en)?|"
        r"ann[eé]e\s+(?:de\s+)?construction)\s*:?\s*(\d{4})",
        text,
        flags=re.I,
    )
    if m:
        y = int(m.group(1))
        if 1700 <= y <= 2100:
            out["annee_construction"] = y

    # Évaluation municipale. Sur Centris : « ÉVALUATION MUNICIPALE
    # (2026) Terrain 48 600 $ Bâtiment 481 400 $ Total 530 000 $ ».
    # On saute le millésime entre parenthèses et on vise le « Total »
    # du bloc (sinon le regex prenait l'année, ou le Terrain).
    m = re.search(
        r"[ée]valuation\s+municipale\s*(?:\(\s*\d{4}\s*\))?"
        r"[\s\S]{0,90}?total[^\d$]{0,6}?([\d][\d\s.,]*\d)\s*\$",
        text,
        flags=re.I,
    )
    if not m:
        m = re.search(
            r"[ée]valuation\s+municipale(?:\s+totale)?\s*"
            r"(?:\(\s*\d{4}\s*\))?[^\d$]{0,8}?([\d][\d\s.,]*\d)\s*\$",
            text,
            flags=re.I,
        )
    if m:
        v = _int_or_none(m.group(1))
        if v and v >= 1_000:
            out["evaluation_municipale"] = v

    # Revenus bruts annuels. Trailing « $ » exigé pour ne capter
    # qu'un montant (« Revenus bruts potentiels 72 600 $ »).
    m = re.search(
        r"revenus?\s*(?:bruts?\s+)?(?:potentiels?\s+)?(?:annuels?\s+)?"
        r"[^\d$]{0,15}?([\d][\d\s.,]*\d)\s*\$",
        text,
        flags=re.I,
    )
    if m:
        v = _int_or_none(m.group(1))
        if v and v >= 1_000:
            out["revenus_bruts"] = v

    # Taxes municipales. Centris affiche le bloc TAXES en deux
    # colonnes — mensuel ET annuel (« Municipales (2026) 760 $ …
    # 9 115 $ »). Le champ vise l'annuel : on capte tous les montants
    # et on retient le plus grand (l'annuel = 12× le mensuel). On
    # saute un millésime entre parenthèses, le « $ » final garantit
    # qu'on prend bien le montant.
    muni_vals = []
    for tm in re.finditer(
        r"taxes?\s+municipal\w*\s*(?:\(\s*\d{4}\s*\))?"
        r"[^\d$]{0,8}?([\d][\d\s.,]*\d)\s*\$",
        text,
        flags=re.I,
    ):
        v = _int_or_none(tm.group(1))
        if v and 100 <= v <= 5_000_000:
            muni_vals.append(v)
    if muni_vals:
        out["taxes_municipales"] = max(muni_vals)

    # Taxes scolaires. Même bloc mensuel/annuel — on garde l'annuel
    # (plus grand montant). « taxes » optionnel : Centris liste juste
    # « Scolaires (2026) 362 $ » sous l'en-tête TAXES.
    sco_vals = []
    for ts in re.finditer(
        r"(?:taxes?\s+)?scolaires?\s*(?:\(\s*\d{4}\s*\))?"
        r"[^\d$]{0,8}?([\d][\d\s.,]*\d)\s*\$",
        text,
        flags=re.I,
    ):
        v = _int_or_none(ts.group(1))
        if v and 50 <= v <= 5_000_000:
            sco_vals.append(v)
    if sco_vals:
        out["taxes_scolaires"] = max(sco_vals)

    # Assurances. Trailing « $ » exigé : évite de capter la pub
    # « assurances auto et habitation » et vise « Assurances 5 258 $ ».
    # Centris affiche les dépenses en deux colonnes mensuel + annuel
    # (comme pour taxes muni/scolaires). On capte tous les montants
    # et on retient le plus grand = l'annuel (= 12× le mensuel). Avant
    # ce fix, re.search prenait le premier match (= mensuel sur
    # Centris moderne) → extraction de 232 $ au lieu de 2 784 $.
    ass_vals = []
    for am in re.finditer(
        r"assurances?[^\d$]{0,12}?([\d][\d\s.,]*\d)\s*\$",
        text,
        flags=re.I,
    ):
        v = _int_or_none(am.group(1))
        if v and 100 <= v <= 5_000_000:
            ass_vals.append(v)
    if ass_vals:
        out["assurances"] = max(ass_vals)

    # Énergie payée par le propriétaire. Sur Centris la section
    # « Dépenses » liste plusieurs postes énergétiques séparés
    # (Électricité, Huile à fournaise, Gaz…) — on les ADDITIONNE
    # tous dans `energie`. Trailing « $ » exigé (vise un montant).
    # Défense en profondeur : pour chaque poste on prend le MAX des
    # matches (au cas où Centris affiche mensuel + annuel par poste,
    # comme c'est déjà le cas pour taxes muni/scolaires et assurances).
    # Si un poste n'a qu'une seule occurrence, max([x]) = x → no-op.
    energie_postes = (
        r"[eé]lectricit[eé]",
        r"hydro(?:[ -]?qu[eé]bec)?",
        r"huile\s+[aà]\s+fournaise",
        r"mazout",
        r"gaz\s+naturel",
        r"gaz",
        r"[eé]nergie",
        r"chauffage",
    )
    energie_total = 0
    for poste in energie_postes:
        poste_vals = []
        for em in re.finditer(
            poste + r"[^\d$]{0,12}?([\d][\d\s.,]*\d)\s*\$",
            text,
            flags=re.I,
        ):
            v = _int_or_none(em.group(1))
            if v and 50 <= v <= 5_000_000:
                poste_vals.append(v)
        if poste_vals:
            energie_total += max(poste_vals)
    if energie_total > 0:
        out["energie"] = energie_total

    # Stationnements.
    m = re.search(
        r"(?:nombre\s+de\s+)?stationnements?[^\d]{0,15}?(\d{1,3})",
        text,
        flags=re.I,
    )
    if m:
        v = int(m.group(1))
        if 0 <= v <= 999:
            out["nb_stationnements"] = v

    # Superficie terrain / bâtiment (en pi² souvent).
    m = re.search(
        r"superficie\s+(?:du\s+)?terrain[^\d]{0,15}?"
        r"([\d][\d\s\.,]*\d)\s*(?:pi|pieds|m)?",
        text,
        flags=re.I,
    )
    if m:
        v = _int_or_none(m.group(1))
        if v and v >= 100:
            out["superficie_terrain"] = v
    m = re.search(
        r"superficie\s+(?:du\s+)?b[aâ]timent[^\d]{0,15}?"
        r"([\d][\d\s\.,]*\d)\s*(?:pi|pieds|m)?",
        text,
        flags=re.I,
    )
    if m:
        v = _int_or_none(m.group(1))
        if v and v >= 100:
            out["superficie_batiment"] = v

    # Code postal canadien (format A1A 1A1).
    m = _POSTAL_RE.search(text)
    if m:
        out["postal_code"] = f"{m.group(1).upper()} {m.group(2).upper()}"

    # Adresse civique « 3715-3737 Rue Ethel », « 1234 Boulevard X »,
    # « 5678 Avenue Y ». On accepte un range numérique optionnel.
    # Types de voies élargis pour couvrir le contexte québécois :
    # Rang (zones agricoles/rurales), Côte (Côte-des-Neiges), Impasse,
    # Allée, Bretelle, Carré, Quai, Promenade, Voie, Cours — en plus
    # des classiques Rue/Boulevard/Avenue/Chemin/Route/Place/Montée/
    # Terrasse/Croissant. Avant ce fix, ces voies ratent le regex et
    # le H1 fautif de parse_centris gagne par défaut.
    m = re.search(
        r"\b(\d{1,5}(?:\s*-\s*\d{1,5})?)[\s,]+"
        r"((?:Rue|Boul(?:evard)?\.?|Av(?:enue)?\.?|Ch(?:emin)?\.?|"
        r"Rte|Route|Place|Pl\.?|Mont[eé]e|Terrasse|Croissant|"
        r"Rang|C[oô]te|Impasse|All[eé]e|Bretelle|Carr[eé]|Quai|"
        r"Promenade|Voie|Cours)\s+"
        r"[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\-' \.]{2,50})",
        text,
        flags=re.I,
    )
    if m:
        addr = f"{m.group(1).strip()} {m.group(2).strip()}"
        out["address"] = re.sub(r"\s+", " ", addr).strip()[:200]

    # Ville. Heuristiques par ordre de fiabilité décroissante :
    city_val: Optional[str] = None
    # Liste noire de "fausses villes" : termes ambigus qui ne sont pas
    # réellement des municipalités (la province, le pays, un libellé
    # générique). Si on tombe sur l'un d'eux, on ignore et on tente
    # une autre heuristique.
    _city_blacklist = {
        "quebec",
        "québec",
        "canada",
        "qc",
        "ontario",
        "on",
        "province",
        "ville",
        "municipalite",
        "municipalité",
    }

    def _accept_city(name: Optional[str]) -> Optional[str]:
        if not name:
            return None
        cleaned = name.strip().strip(".,;:").strip()
        if not cleaned or len(cleaned) < 2:
            return None
        if cleaned.lower() in _city_blacklist:
            return None
        return cleaned

    #   a) format Centris « …, <Ville> - Ville » (ex. Valcourt - Ville)
    m = re.search(
        r",\s*([A-ZÀ-Ÿ][A-Za-zÀ-ÿ\-']+(?:\s[A-ZÀ-Ÿ][A-Za-zÀ-ÿ\-']+)?)"
        r"\s*-\s*Ville\b",
        text,
    )
    if m:
        city_val = _accept_city(m.group(1))
    #   b) format québécois canonique « <adresse>, <Ville> (QC) »
    if not city_val:
        m = re.search(
            r",\s*([A-ZÀ-Ÿ][A-Za-zÀ-ÿ\-' ]{2,40}?)\s*"
            r"\(\s*(?:QC|Qu[eé]bec|ON|Ontario)\s*\)",
            text,
        )
        if m:
            city_val = _accept_city(m.group(1))
    #   c) format SANS « : » ni virgule préalable :
    #      « <Ville> (QC) <CP> » ou « <Ville> (QC) » en tête de ligne /
    #      après un espace. Utile pour les courriels non structurés
    #      ou les textes Centris où l'adresse précédente est absente.
    if not city_val:
        m = re.search(
            r"(?:^|[\s>])([A-ZÀ-Ÿ][A-Za-zÀ-ÿ\-' ]{2,40}?)\s*"
            r"\(\s*(?:QC|Qu[eé]bec|ON|Ontario)\s*\)"
            r"(?:\s+[A-Z]\d[A-Z]\s?\d[A-Z]\d)?",
            text,
        )
        if m:
            city_val = _accept_city(m.group(1))
    #   d) « Ville : <Nom> » / « Municipalité : <Nom> ». Le « : » est
    #      EXIGÉ pour ce pattern : ça couvre les courriels structurés
    #      type « Ville : Montréal », « Municipalité : Sherbrooke ».
    if not city_val:
        m = re.search(
            r"(?:Ville|Municipalit[eé])\s*:\s*"
            r"([A-ZÀ-Ÿ][A-Za-zÀ-ÿ\-' ]{2,40})",
            text,
        )
        if m:
            city_val = _accept_city(m.group(1))
    if city_val:
        out["city"] = city_val

    # Province (presque toujours QC, mais parfois explicite).
    if re.search(r"\bQu[eé]bec\b|\bQC\b", text):
        out["province"] = "QC"

    # Type de bâtiment.
    for kw, label in (
        (r"\btriplex\b", "Triplex"),
        (r"\bduplex\b", "Duplex"),
        (r"\bquadruplex\b", "Quadruplex"),
        (r"\bquintuplex\b", "Quintuplex"),
        (r"\bplex\b", "Plex"),
        (r"\bcondominium\b|\bcondo\b", "Condo"),
        (r"\bmulti(?:logements?|plex)?\b", "Multilogements"),
        (r"\bmixte\b", "Mixte"),
    ):
        if re.search(kw, text, flags=re.I):
            out["type_batiment"] = label
            break

    # Courtier — nom probable. On combine re.I sur le préfixe MAIS on
    # reste case-sensitive sur le nom (sinon le nom est confondu avec
    # le mot-clé) : deux étapes — trouver le préfixe, matcher le nom.
    # Priorité 1 : la section Centris « Courtiers inscripteurs » suivie
    # directement du nom de l'inscripteur. Priorité 2 : « Courtier : ».
    _NAME_RE = (
        r"([A-ZÀ-Ÿ][a-zà-ÿ\-']+(?:\s+[A-ZÀ-Ÿ][a-zà-ÿ\-']+){1,3})"
    )
    pref = re.search(
        r"courtiers?\s+inscripteurs?\s*",
        text,
        flags=re.I,
    )
    if not pref:
        pref = re.search(
            r"\b(?:courtier|agent|repr[eé]sentant)\b\s*:?\s*",
            text,
            flags=re.I,
        )
    if pref:
        tail = text[pref.end():pref.end() + 120]
        nm = re.match(_NAME_RE, tail)
        if nm:
            out["courtier_nom"] = nm.group(1).strip()

    # Courtier — téléphone + email regroupés.
    contact_bits: List[str] = []
    m = _PHONE_RE.search(text)
    if m:
        contact_bits.append(f"{m.group(1)}-{m.group(2)}-{m.group(3)}")
    m = _EMAIL_RE.search(text)
    if m:
        contact_bits.append(m.group(0))
    if contact_bits:
        out["courtier_contact"] = " / ".join(contact_bits)

    return out


# ── Parser JSON-LD (schema.org) ───────────────────────────────────


_JSONLD_RE = re.compile(
    r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>'
    r'([\s\S]*?)</script>',
    flags=re.I,
)


_SCHEMA_TYPES = {
    "product",
    "realestatelisting",
    "house",
    "apartment",
    "singlefamilyresidence",
    "residence",
    "place",
    "accommodation",
    "offer",
}


def _walk_jsonld(node: Any, found: List[dict]) -> None:
    """Aplatit un graphe JSON-LD pour récupérer tous les sous-objets
    typés (schema.org). Robuste aux @graph, listes imbriquées, etc."""
    if isinstance(node, dict):
        t = node.get("@type")
        types = t if isinstance(t, list) else [t] if t else []
        if any(
            isinstance(x, str) and x.lower() in _SCHEMA_TYPES
            for x in types
        ):
            found.append(node)
        for v in node.values():
            _walk_jsonld(v, found)
    elif isinstance(node, list):
        for v in node:
            _walk_jsonld(v, found)


def parse_jsonld(html: str) -> Dict[str, Any]:
    """Cherche les blocs ``<script type="application/ld+json">`` et en
    extrait les champs immobiliers connus de schema.org."""
    out: Dict[str, Any] = {}
    for raw in _JSONLD_RE.findall(html):
        try:
            parsed = json.loads(raw.strip())
        except (ValueError, TypeError):
            continue
        found: List[dict] = []
        _walk_jsonld(parsed, found)
        for node in found:
            _merge_jsonld_node(node, out)
    return out


def _merge_jsonld_node(node: dict, out: dict) -> None:
    """Extrait les champs reconnus d'un nœud schema.org et les pose
    dans ``out`` s'ils ne sont pas déjà présents."""

    def _put(key: str, value: Any) -> None:
        if value is None or value == "" or out.get(key) is not None:
            return
        out[key] = value

    # Nom / description
    _put("description", node.get("description") or node.get("name"))

    # Address (peut être string ou objet PostalAddress).
    addr = node.get("address")
    if isinstance(addr, dict):
        street = addr.get("streetAddress")
        city = addr.get("addressLocality")
        postal = addr.get("postalCode")
        prov = addr.get("addressRegion")
        if street:
            _put("address", str(street).strip()[:200])
        if city:
            _put("city", str(city).strip()[:100])
        if postal:
            _put("postal_code", str(postal).strip().upper()[:10])
        if prov:
            _put("province", str(prov).strip().upper()[:5])
    elif isinstance(addr, str):
        _put("address", addr.strip()[:200])

    # Prix : offers.price, offers.priceSpecification.price, price.
    offers = node.get("offers")
    if isinstance(offers, dict):
        price = offers.get("price") or offers.get("lowPrice")
        if price is None:
            ps = offers.get("priceSpecification")
            if isinstance(ps, dict):
                price = ps.get("price")
        v = _int_or_none(price)
        if v:
            _put("asking_price", v)
    elif isinstance(offers, list) and offers:
        v = _int_or_none(offers[0].get("price"))
        if v:
            _put("asking_price", v)
    else:
        v = _int_or_none(node.get("price"))
        if v:
            _put("asking_price", v)

    # Année construction (yearBuilt / dateBuilt).
    y = _int_or_none(node.get("yearBuilt") or node.get("dateBuilt"))
    if y and 1700 <= y <= 2100:
        _put("annee_construction", y)

    # Nb d'unités (numberOfRooms est ambigu — on l'évite ici).
    v = _int_or_none(node.get("numberOfUnits"))
    if v:
        _put("nb_logements", v)

    # Surfaces.
    fs = node.get("floorSize")
    if isinstance(fs, dict):
        v = _int_or_none(fs.get("value"))
        if v:
            _put("superficie_batiment", v)
    elif fs is not None:
        v = _int_or_none(fs)
        if v:
            _put("superficie_batiment", v)

    ls = node.get("lotSize")
    if isinstance(ls, dict):
        v = _int_or_none(ls.get("value"))
        if v:
            _put("superficie_terrain", v)
    elif ls is not None:
        v = _int_or_none(ls)
        if v:
            _put("superficie_terrain", v)


# ── Parser __NEXT_DATA__ (Next.js — pmml.ca, etc.) ────────────────


_NEXTDATA_RE = re.compile(
    r'<script[^>]*id=["\']__NEXT_DATA__["\'][^>]*>([\s\S]*?)</script>',
    flags=re.I,
)


def has_next_data(html: str) -> bool:
    return bool(_NEXTDATA_RE.search(html))


def parse_next_data(html: str) -> Dict[str, Any]:
    """Extrait depuis le bloc ``__NEXT_DATA__`` (Next.js). Les sites
    immobiliers Next.js (pmml.ca par ex.) sérialisent leur état dans
    ``props.pageProps`` — on cherche récursivement les champs connus
    (price, address, units, etc.).

    Bonus : on convertit aussi l'arbre JSON en texte stringifié et on
    le passe à ``parse_text`` — ça capte les champs québécois rares
    (typologie, taxes, évaluation) que le SPA stocke en chaînes
    libres dans le JSON."""
    m = _NEXTDATA_RE.search(html)
    if not m:
        return {}
    try:
        parsed = json.loads(m.group(1).strip())
    except (ValueError, TypeError):
        return {}

    out: Dict[str, Any] = {}
    _walk_next_data(parsed, out)

    # Filet de sécurité : passer le JSON stringifié à parse_text pour
    # rattraper ce qu'on aurait raté. Les valeurs déjà extraites par
    # le walker structurel ne sont pas écrasées.
    try:
        text_blob = json.dumps(parsed, ensure_ascii=False)
        text_out = parse_text(text_blob)
        for k, v in text_out.items():
            out.setdefault(k, v)
    except (TypeError, ValueError):
        pass

    return out


# Clés JSON couramment utilisées par les SPA immobiliers (Next.js
# notamment) → mapping vers nos champs canoniques.
_NEXT_FIELD_MAP: Dict[str, str] = {
    "price": "asking_price",
    "askingprice": "asking_price",
    "listprice": "asking_price",
    "salesprice": "asking_price",
    "prixdemande": "asking_price",
    "address": "address",
    "streetaddress": "address",
    "adresse": "address",
    "city": "city",
    "ville": "city",
    "postalcode": "postal_code",
    "codepostal": "postal_code",
    "province": "province",
    "yearbuilt": "annee_construction",
    "anneeconstruction": "annee_construction",
    "numberofunits": "nb_logements",
    "nbunits": "nb_logements",
    "nblogements": "nb_logements",
    "units": "nb_logements",
    "description": "description",
    "evaluationmunicipale": "evaluation_municipale",
    "revenusbruts": "revenus_bruts",
    "taxesmunicipales": "taxes_municipales",
    "taxesscolaires": "taxes_scolaires",
    "lotsize": "superficie_terrain",
    "floorsize": "superficie_batiment",
    "areaterrain": "superficie_terrain",
    "areabatiment": "superficie_batiment",
}


def _walk_next_data(node: Any, out: dict, _depth: int = 0) -> None:
    """Parcourt récursivement un arbre JSON et pose dans ``out`` les
    valeurs des clés connues."""
    if _depth > 25:  # garde-fou contre les cycles improbables
        return
    if isinstance(node, dict):
        for k, v in node.items():
            key_norm = re.sub(r"[^a-z]", "", str(k).lower())
            mapped = _NEXT_FIELD_MAP.get(key_norm)
            if mapped and mapped not in out:
                if mapped in (
                    "asking_price",
                    "annee_construction",
                    "nb_logements",
                    "evaluation_municipale",
                    "revenus_bruts",
                    "taxes_municipales",
                    "taxes_scolaires",
                    "superficie_terrain",
                    "superficie_batiment",
                ):
                    iv = _int_or_none(v)
                    if iv:
                        out[mapped] = iv
                elif isinstance(v, str) and v.strip():
                    out[mapped] = v.strip()[:500]
            _walk_next_data(v, out, _depth + 1)
    elif isinstance(node, list):
        for v in node:
            _walk_next_data(v, out, _depth + 1)


# ── Parsers spécifiques sites ─────────────────────────────────────


def _get_soup(html: str):
    """Crée un BeautifulSoup en gérant l'import paresseusement (le
    paquet est lourd à importer ; sur des URLs sans HTML utile on
    veut éviter le coût)."""
    from bs4 import BeautifulSoup  # type: ignore

    return BeautifulSoup(html, "html.parser")


def _strip_html(html: str) -> str:
    """Strip HTML → texte plat. Utilisé en fallback regex."""
    text = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
    text = re.sub(r"<!--[\s\S]*?-->", " ", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"&amp;", "&", text)
    text = re.sub(r"&#x27;|&#39;", "'", text)
    text = re.sub(r"&quot;", '"', text)
    text = re.sub(r"&#?\w+;", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_centris(html: str) -> Dict[str, Any]:
    """Parser sélecteurs CSS spécifiques à Centris.ca. En cas d'échec
    sur un sélecteur, on retombe gracieusement sur JSON-LD puis sur
    parse_text(HTML stripped)."""
    out: Dict[str, Any] = {}
    try:
        soup = _get_soup(html)
    except Exception:  # noqa: BLE001
        return parse_text(_strip_html(html))

    # Prix : Centris utilise .priceTag, ou un <span itemprop="price">,
    # ou un <meta itemprop="price">.
    for sel in (".priceTag", "[itemprop='price']", "meta[itemprop='price']"):
        node = soup.select_one(sel)
        if node:
            raw = node.get("content") or node.get_text(" ", strip=True)
            v = _int_or_none(raw)
            if v and v >= 10_000:
                out["asking_price"] = v
                break

    # Adresse : .address ou h1. Filtre négatif appliqué : sur Centris
    # moderne, le H1 affiche souvent le TYPE de propriété (« Immeuble
    # à revenu à vendre ») au lieu de l'adresse civique. On rejette
    # toute valeur qui ne contient pas de chiffre OU qui contient un
    # terme invalidant (immeuble, plex, à vendre, etc.). En cas de
    # rejet sur tous les sélecteurs, parse_text aura sa chance via le
    # fallback regex plus bas.
    _addr_invalid = re.compile(
        r"\b(?:[aà]\s+vendre|[aà]\s+louer|immeuble|propri[eé]t[eé]|"
        r"plex|duplex|triplex|quadruplex|quintuplex|"
        r"r[eé]sidentiel|commercial|condominium|condo|"
        r"multilogements?|mixte)\b",
        re.I,
    )
    for sel in (".address", "h1[itemprop='address']", "h1.heading"):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt:
                # Validation : doit contenir au moins un chiffre ET
                # ne contenir aucun terme invalidant.
                if (not re.search(r"\d", txt)) or _addr_invalid.search(txt):
                    log.debug(
                        "parse_centris: address selector %r rejected "
                        "(invalid value: %r)",
                        sel,
                        txt[:120],
                    )
                    continue
                # Souvent Centris cumule "adresse, ville (province)" → on
                # split sur la 1re virgule.
                if "," in txt:
                    parts = [p.strip() for p in txt.split(",", 1)]
                    out["address"] = parts[0][:200]
                    if len(parts) == 2:
                        rest = parts[1]
                        # « Verdun (Montréal) (QC) H4G 1M4 »
                        city_m = re.match(
                            r"([^\(]+)\s*(?:\([^\)]+\))?\s*\(([A-Z]{2})\)?",
                            rest,
                        )
                        if city_m:
                            out["city"] = city_m.group(1).strip()
                            out["province"] = city_m.group(2)
                        pm = _POSTAL_RE.search(rest)
                        if pm:
                            out["postal_code"] = (
                                f"{pm.group(1).upper()} {pm.group(2).upper()}"
                            )
                else:
                    out["address"] = txt[:200]
                break

    # Description (.description, .summary, .text).
    for sel in (".description", ".summary", "[itemprop='description']"):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt and len(txt) > 30:
                out["description"] = txt[:2000]
                break

    # Ville : sélecteurs CSS dédiés (Centris breadcrumb / itemprop /
    # data-id). Appliqués UNIQUEMENT si la ville n'a pas été captée
    # via l'address parsing plus haut. Sinon JSON-LD prendra le relais.
    if "city" not in out:
        _city_blacklist_html = {
            "quebec", "québec", "canada", "qc", "ontario", "on",
            "province", "ville", "municipalité", "municipalite",
            "accueil", "home",
        }
        for sel in (
            "[itemprop='addressLocality']",
            "nav[aria-label*='breadcrumb' i] a:last-child",
            "nav[aria-label*='fil' i] a:last-child",
            ".breadcrumb li:last-child a",
            ".breadcrumb li:last-child",
            ".bread-crumb-item:last-child",
            "[data-id*='Locality']",
        ):
            try:
                node = soup.select_one(sel)
            except Exception:  # noqa: BLE001 — sélecteur invalide
                continue
            if not node:
                continue
            txt = node.get_text(" ", strip=True)
            if not txt or len(txt) < 2 or len(txt) > 60:
                continue
            if txt.lower() in _city_blacklist_html:
                continue
            # Doit ressembler à un nom propre (commence par majuscule
            # ou caractère accentué). Évite "à vendre" et autres.
            if not re.match(r"^[A-ZÀ-Ÿ]", txt):
                continue
            out["city"] = txt[:100]
            break

    # Code postal : sélecteur CSS dédié, si présent.
    if "postal_code" not in out:
        for sel in ("[itemprop='postalCode']", ".postal-code", ".postalCode"):
            try:
                node = soup.select_one(sel)
            except Exception:  # noqa: BLE001
                continue
            if not node:
                continue
            raw = node.get("content") or node.get_text(" ", strip=True)
            if not raw:
                continue
            pm = _POSTAL_RE.search(raw)
            if pm:
                out["postal_code"] = (
                    f"{pm.group(1).upper()} {pm.group(2).upper()}"
                )
                break

    # Fallback massif sur JSON-LD pour les champs encore manquants
    # (address.addressLocality, address.postalCode pris en charge par
    # parse_jsonld → _merge_jsonld_node).
    ld = parse_jsonld(html)
    for k, v in ld.items():
        out.setdefault(k, v)

    # Fallback regex sur le texte strippé pour ce qui reste (taxes,
    # typologie, année, etc.).
    text_out = parse_text(_strip_html(html))
    for k, v in text_out.items():
        out.setdefault(k, v)

    return out


def parse_duproprio(html: str) -> Dict[str, Any]:
    """Parser DuProprio.com. Approche similaire à Centris : sélecteurs
    courants + fallback JSON-LD + fallback regex texte."""
    out: Dict[str, Any] = {}
    try:
        soup = _get_soup(html)
    except Exception:  # noqa: BLE001
        return parse_text(_strip_html(html))

    # Prix demandé : sur DuProprio, classes possibles : .listing-price,
    # .price__value, [data-qaid="listing-price"].
    for sel in (
        ".listing-price",
        ".price__value",
        "[data-qaid='listing-price']",
        ".sc-price",
        "meta[itemprop='price']",
    ):
        node = soup.select_one(sel)
        if node:
            raw = node.get("content") or node.get_text(" ", strip=True)
            v = _int_or_none(raw)
            if v and v >= 10_000:
                out["asking_price"] = v
                break

    # Adresse.
    for sel in (
        ".listing-location__address",
        "h1.listing-title",
        "[data-qaid='listing-address']",
        "h1",
    ):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt:
                out.setdefault("address", txt[:200])
                break

    # Description.
    for sel in (
        ".listing-description",
        "[data-qaid='listing-description']",
        ".sc-description",
    ):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt and len(txt) > 30:
                out["description"] = txt[:2000]
                break

    ld = parse_jsonld(html)
    for k, v in ld.items():
        out.setdefault(k, v)
    text_out = parse_text(_strip_html(html))
    for k, v in text_out.items():
        out.setdefault(k, v)
    return out


def parse_realtor(html: str) -> Dict[str, Any]:
    """Parser Realtor.ca. Realtor.ca est très protégé (Cloudflare), le
    fetch HTTP direct retourne souvent du HTML minimal — on tape donc
    JSON-LD si présent, et le reste en regex texte."""
    out: Dict[str, Any] = {}
    try:
        soup = _get_soup(html)
    except Exception:  # noqa: BLE001
        return parse_text(_strip_html(html))

    # Prix : #listingPriceValue, #priceColLeft.
    for sel in (
        "#listingPriceValue",
        "#priceColLeft",
        "[itemprop='price']",
        "meta[itemprop='price']",
    ):
        node = soup.select_one(sel)
        if node:
            raw = node.get("content") or node.get_text(" ", strip=True)
            v = _int_or_none(raw)
            if v and v >= 10_000:
                out["asking_price"] = v
                break

    # Adresse.
    for sel in (
        "#listingAddress",
        "h1[itemprop='address']",
        ".listingAddress",
    ):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt:
                out.setdefault("address", txt[:200])
                break

    # Description.
    for sel in (
        "#propertyDescriptionCon",
        "[itemprop='description']",
        ".propertyDescription",
    ):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            if txt and len(txt) > 30:
                out["description"] = txt[:2000]
                break

    ld = parse_jsonld(html)
    for k, v in ld.items():
        out.setdefault(k, v)
    text_out = parse_text(_strip_html(html))
    for k, v in text_out.items():
        out.setdefault(k, v)
    return out


# ── Parser PDF (pypdf) ────────────────────────────────────────────


def parse_pdf(blob: bytes) -> str:
    """Extrait le texte brut d'un PDF (toutes pages concaténées).
    Retourne une chaîne vide si le PDF est purement scanné (pas de
    couche texte) — le caller émettra alors un warning."""
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        log.warning("pypdf non installé — extraction PDF désactivée")
        return ""
    try:
        reader = PdfReader(io.BytesIO(blob))
        chunks: List[str] = []
        for page in reader.pages:
            try:
                t = page.extract_text() or ""
            except Exception:  # noqa: BLE001
                t = ""
            if t.strip():
                chunks.append(t)
        return "\n".join(chunks)
    except Exception as exc:  # noqa: BLE001
        log.warning("pypdf extraction failed: %s", exc)
        return ""


# ── Parser Excel (.xlsx / .xls) ───────────────────────────────────
#
# Phil reçoit régulièrement des tableaux Excel (listes d'immeubles à
# vendre, exports de courtiers, listings privés) — souvent une ligne
# par immeuble avec des colonnes type « Adresse / Prix / Revenus /
# Taxes ». On convertit le tableau en texte structuré (headers
# détectés + lignes séparées par « | ») et on le passe à parse_text
# ET à Gemini en parallèle, comme n'importe quel autre input texte.


def _excel_row_text(row_values: List[Any]) -> str:
    """Sérialise une ligne Excel en chaîne « v1 | v2 | v3 »."""
    cells: List[str] = []
    for v in row_values:
        if v is None:
            cells.append("")
            continue
        if isinstance(v, float) and v.is_integer():
            cells.append(str(int(v)))
        else:
            cells.append(str(v).strip())
    return " | ".join(cells)


def _looks_like_header(row_values: List[Any]) -> bool:
    """Heuristique : est-ce que la première ligne ressemble à des
    en-têtes ? On considère que oui si une majorité des cellules sont
    des chaînes courtes contenant un libellé typique (adresse, prix,
    ville, taxes, revenus, etc.). Permet de mieux mapper les colonnes
    plus tard et d'enrichir le texte structuré passé aux parseurs."""
    if not row_values:
        return False
    header_kw = {
        "adresse", "address", "ville", "city", "prix", "price",
        "asking", "demande", "demandé", "logements", "unités",
        "unites", "units", "taxes", "revenu", "revenus", "revenue",
        "code postal", "postal", "année", "annee", "construction",
        "superficie", "évaluation", "evaluation", "assurance",
        "énergie", "energie", "type", "courtier",
    }
    matched = 0
    str_cells = 0
    for v in row_values:
        if not isinstance(v, str):
            continue
        str_cells += 1
        low = v.strip().lower()
        if not low:
            continue
        if any(kw in low for kw in header_kw):
            matched += 1
    # Au moins 2 cellules reconnues comme libellés ET majoritairement
    # textuelles → on considère que c'est une ligne d'en-têtes.
    return matched >= 2 and str_cells >= max(2, len(row_values) // 2)


def parse_excel(blob: bytes, filename: str = "excel") -> str:
    """Convertit un fichier Excel en texte structuré exploitable par
    parse_text + Gemini.

    Stratégie :
      - Lit toutes les feuilles non vides via openpyxl (data_only=True
        pour récupérer les valeurs calculées, pas les formules).
      - Si la première ligne semble être des en-têtes (libellés
        reconnus type « Adresse », « Prix »), on les conserve en tête
        du texte et on répète les valeurs sous forme « Header: valeur »
        ligne par ligne — meilleur ancrage pour les regex de
        parse_text et pour la compréhension Gemini.
      - Sinon, on dump simplement les lignes en « v1 | v2 | v3 ».

    Pour .xls (ancien format binaire OLE) on tente xlrd ; si absent
    on retourne une chaîne vide et un caller émet un warning.

    Retourne une chaîne vide en cas d'échec — jamais d'exception.
    """
    is_xls = filename.lower().endswith(".xls")
    t0 = time.perf_counter()

    if is_xls:
        # Ancien format binaire — xlrd ne fait que .xls depuis sa
        # v2.0.1 (Microsoft ayant supprimé le support .xlsx). On
        # essaie mais le paquet n'est pas garanti installé.
        try:
            import xlrd  # type: ignore
        except ImportError:
            log.warning(
                "Excel '%s' : .xls non supporté (xlrd absent). "
                "Phil peut sauver en .xlsx pour traiter ce fichier.",
                filename,
            )
            return ""
        try:
            wb = xlrd.open_workbook(file_contents=blob)
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "Excel '%s' (xls) : ouverture impossible — %s",
                filename,
                exc,
            )
            return ""
        parts: List[str] = []
        for sheet in wb.sheets():
            if sheet.nrows == 0:
                continue
            parts.append(f"[Feuille : {sheet.name}]")
            rows = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            parts.extend(_render_excel_rows(rows))
        text = "\n".join(parts)
        dt = time.perf_counter() - t0
        log.info(
            "Excel xls '%s' : %d feuilles, %d chars en %.2fs",
            filename,
            len(wb.sheets()),
            len(text),
            dt,
        )
        return text

    # .xlsx — format standard depuis Excel 2007.
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        log.warning(
            "Excel '%s' : openpyxl absent — installer pour activer "
            "le support .xlsx",
            filename,
        )
        return ""
    try:
        wb = load_workbook(
            filename=io.BytesIO(blob), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001
        log.warning(
            "Excel '%s' : ouverture openpyxl impossible — %s",
            filename,
            exc,
        )
        return ""

    parts: List[str] = []
    nb_sheets = 0
    for ws in wb.worksheets:
        nb_sheets += 1
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            # Ignore les lignes 100% vides.
            if not any(c not in (None, "") for c in row):
                continue
            rows.append(list(row))
        if not rows:
            continue
        parts.append(f"[Feuille : {ws.title}]")
        parts.extend(_render_excel_rows(rows))
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass

    text = "\n".join(parts)
    dt = time.perf_counter() - t0
    log.info(
        "Excel xlsx '%s' : %d feuilles, %d chars en %.2fs",
        filename,
        nb_sheets,
        len(text),
        dt,
    )
    return text


def _render_excel_rows(rows: List[List[Any]]) -> List[str]:
    """Rend une liste de lignes Excel en texte exploitable.

    Si la première ligne ressemble à des en-têtes, on l'utilise pour
    générer des paires « Libellé: valeur » par cellule (en plus du
    dump tabulaire) — ancrage idéal pour les regex de parse_text.
    """
    out: List[str] = []
    headers: List[str] = []
    data_start = 0
    if _looks_like_header(rows[0]):
        headers = [
            (str(c).strip() if c is not None else "") for c in rows[0]
        ]
        out.append("[En-têtes] " + _excel_row_text(rows[0]))
        data_start = 1

    for row in rows[data_start:]:
        out.append(_excel_row_text(row))
        if headers:
            # On enrichit chaque ligne avec des paires libellé/valeur
            # — parse_text accroche bien sur ce format.
            pairs: List[str] = []
            for i, v in enumerate(row):
                if i >= len(headers):
                    break
                h = headers[i]
                if not h:
                    continue
                if v is None or v == "":
                    continue
                if isinstance(v, float) and v.is_integer():
                    val_s = str(int(v))
                else:
                    val_s = str(v).strip()
                pairs.append(f"{h}: {val_s}")
            if pairs:
                out.append("  " + " ; ".join(pairs))
    return out


# ── Fusion multi-sources ──────────────────────────────────────────


# Priorité décroissante : parser dédié (Centris/DuProprio/Realtor/Next)
# l'emporte sur JSON-LD générique, qui l'emporte sur regex texte.
# Les sources OCR (image-ocr, pdf-ocr) sont notées un cran plus bas que
# le texte natif — Tesseract introduit du bruit, donc si on a un PDF
# avec couche texte ET une image OCR pour le même immeuble, on garde
# les valeurs natives.
_PARSER_PRIORITY: Dict[str, int] = {
    # L'analyse texte de la page Centris (ancrée sur les libellés
    # stables) prime sur le parser CSS, dont les sélecteurs se
    # périment et renvoient parfois des valeurs fausses.
    "centris-text": 105,
    "centris": 100,
    "duproprio": 100,
    "realtor": 100,
    "next_data": 90,
    "jsonld": 70,
    "text": 50,
    "pdf": 50,
    "image-ocr": 40,
    "pdf-ocr": 40,
}


def merge_extracted(sources: List[Tuple[str, Dict[str, Any]]]) -> Dict[str, Any]:
    """Fusionne plusieurs dicts en respectant la priorité du parseur.

    ``sources`` est une liste de ``(parser_tag, dict)``. Pour chaque
    champ, on garde la valeur du parseur le plus prioritaire qui a
    fourni une valeur non-nulle. À priorité égale, premier arrivé.
    La typologie est fusionnée champ par champ (les sources peuvent
    se compléter).
    """
    merged: Dict[str, Any] = {}
    field_origin: Dict[str, int] = {}

    for tag, data in sources:
        if not data:
            continue
        prio = _PARSER_PRIORITY.get(tag, 0)
        for k, v in data.items():
            if v is None or v == "":
                continue
            if k == "typology" and isinstance(v, dict):
                existing = merged.get("typology") or {}
                for tk, tv in v.items():
                    # Pour la typologie : on prend le MAX (le texte cite
                    # la typo plusieurs fois, on veut le compte réel).
                    existing[tk] = max(existing.get(tk, 0), int(tv))
                merged["typology"] = existing
                field_origin[k] = max(field_origin.get(k, 0), prio)
                continue
            if k not in merged or prio > field_origin.get(k, -1):
                merged[k] = v
                field_origin[k] = prio
    return merged


# ── Fetch URL ─────────────────────────────────────────────────────


async def _fetch_html(url: str) -> Tuple[str, Optional[str]]:
    """Télécharge le HTML d'une URL. Retourne (html, error_msg).
    En cas d'échec, html est "" et error_msg explique pourquoi."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "fr-CA,fr;q=0.9,en;q=0.8",
    }
    try:
        async with httpx.AsyncClient(
            timeout=20.0, follow_redirects=True
        ) as client:
            r = await client.get(url, headers=headers)
            if r.status_code >= 400:
                return "", (
                    f"page inaccessible (HTTP {r.status_code}) — "
                    "le site bloque peut-être les fetchs automatisés"
                )
            return r.text, None
    except httpx.TimeoutException:
        return "", "timeout après 20 s"
    except Exception as exc:  # noqa: BLE001
        return "", f"erreur réseau : {exc!s}"


async def _fetch_html_rendered(
    url: str, parser_kind: str
) -> Tuple[str, Optional[str]]:
    """Comme ``_fetch_html`` mais, pour Centris, télécharge le HTML
    *rendu* (JavaScript exécuté) via le VPS Playwright quand il est
    configuré.

    Centris injecte la section « Détails financiers » (évaluation
    municipale, taxes municipales/scolaires, dépenses/énergie) en
    JavaScript : un fetch httpx direct ne récupère pas ce bloc. Le
    navigateur Chromium du VPS exécute le JS, déplie les sections et
    renvoie le DOM complet — les parsers texte ancrés sur les
    libellés y retrouvent alors taxes/évaluation/dépenses.

    Fallback transparent sur le fetch httpx direct si le VPS n'est
    pas configuré, ne renvoie pas de HTML, ou échoue.
    """
    if parser_kind == "centris" and scraping_proxy.vps_available():
        try:
            detail = await scraping_proxy.scrape_centris_detail(url)
            html = (detail or {}).get("html") or ""
            if html.strip():
                log.info(
                    "Centris %s : HTML rendu via VPS (%d chars)",
                    url,
                    len(html),
                )
                return html, None
            log.warning(
                "VPS Centris : pas de HTML rendu pour %s — "
                "fallback httpx direct",
                url,
            )
        except Exception as exc:  # noqa: BLE001
            log.warning(
                "VPS Centris a échoué pour %s (%s) — fallback httpx",
                url,
                exc,
            )
    return await _fetch_html(url)


def _parser_for_domain(domain: str) -> str:
    d = domain.lower()
    if "centris.ca" in d:
        return "centris"
    if "duproprio.com" in d:
        return "duproprio"
    if "realtor.ca" in d:
        return "realtor"
    return "generic"


# ── Merge intelligent local + Gemini ──────────────────────────────


def _norm_str(s: Any) -> str:
    """Normalise une chaîne pour comparaison : trim, casse repliée,
    espaces multiples écrasés, ponctuation finale enlevée."""
    if s is None:
        return ""
    out = str(s).strip().lower()
    out = re.sub(r"\s+", " ", out)
    out = out.strip(".,;:")
    return out


def _values_concordant(a: Any, b: Any, field_name: str) -> bool:
    """Retourne True si deux valeurs sont jugées concordantes.

    - Numériques (montants, années, surfaces, nb logements…) : à
      moins de 5% l'une de l'autre (ou exactement égales pour les
      très petits nombres < 20, où le delta absolu prime).
    - Typology (dict) : mêmes clés et mêmes valeurs (ordre indifférent).
    - Strings : égalité après normalisation.
    """
    if a is None or a == "":
        return False
    if b is None or b == "":
        return False
    if field_name == "typology":
        if not isinstance(a, dict) or not isinstance(b, dict):
            return False
        if set(a.keys()) != set(b.keys()):
            return False
        for k, v in a.items():
            try:
                if int(v) != int(b.get(k, -1)):
                    return False
            except (TypeError, ValueError):
                if v != b.get(k):
                    return False
        return True
    if field_name in _NUMERIC_FIELDS:
        va = normalize_number(a)
        vb = normalize_number(b)
        if va is None or vb is None:
            return False
        if va == 0 and vb == 0:
            return True
        if va < 20 and vb < 20:
            return abs(va - vb) < 0.5
        denom = max(abs(va), abs(vb))
        return abs(va - vb) / denom <= _NUMERIC_TOLERANCE
    # Strings et autres.
    return _norm_str(a) == _norm_str(b)


def _looks_like_valid_address(s: Any) -> bool:
    """Heuristique : adresse civique « clairement valide » côté
    parser local. Commence par un chiffre (numéro civique) et
    contient au moins un mot de 3+ lettres."""
    if not s:
        return False
    raw = str(s).strip()
    if not raw or not raw[0].isdigit():
        return False
    return bool(re.search(r"[A-Za-zÀ-ÿ]{3,}", raw))


def _merge_local_gemini(
    local: Dict[str, Any],
    gemini: Dict[str, Any],
) -> Tuple[Dict[str, Any], List[str], int]:
    """Fusionne un dict du parser local et un dict Gemini selon la
    matrice de décision spec Phase A1 (cf. docstring module).

    Retourne ``(merged, divergence_warnings, divergences_count)``.
    Les warnings émis sont visibles côté UI ; ils citent le champ,
    les deux valeurs concurrentes et la décision prise.
    """
    merged: Dict[str, Any] = {}
    divergence_warnings: List[str] = []
    divergences = 0

    all_keys = set(local.keys()) | set(gemini.keys())
    for k in all_keys:
        v_local = local.get(k)
        v_gem = gemini.get(k)
        has_local = v_local not in (None, "", {}, [])
        has_gem = v_gem not in (None, "", {}, [])

        if has_local and has_gem:
            if _values_concordant(v_local, v_gem, k):
                # Concordance → confiance forte, on garde local
                # (parsers spécifiques Centris/DuProprio/Realtor
                # sont fiables sur les valeurs qu'ils sortent).
                merged[k] = v_local
            elif k == "typology" and isinstance(v_local, dict) and isinstance(v_gem, dict):
                # Pour la typologie en divergence, on union les
                # clés en prenant le MAX par typologie — chaque
                # source peut compléter l'autre sans perte.
                combined = dict(v_local)
                for tk, tv in v_gem.items():
                    try:
                        combined[tk] = max(
                            int(combined.get(tk, 0)), int(tv)
                        )
                    except (TypeError, ValueError):
                        combined.setdefault(tk, tv)
                merged[k] = combined
            elif k in _GEO_FIELDS and _looks_like_valid_address(v_local):
                # Adresse locale clairement valide → on la garde
                # malgré la divergence (Gemini hallucine parfois
                # sur l'adresse civique).
                merged[k] = v_local
                divergences += 1
                divergence_warnings.append(
                    f"Divergence sur {k} : parser local = "
                    f"« {v_local} », Gemini = « {v_gem} ». Pris "
                    "parser local (adresse civique valide). "
                    "Vérifier manuellement."
                )
            elif k in _GEO_FIELDS and k != "address":
                # Pour city/postal_code/province : Gemini est
                # souvent meilleur sur l'inférence (déduit la
                # ville à partir d'autres indices) — sauf si le
                # local a une valeur visiblement bonne.
                if k == "postal_code" and re.fullmatch(
                    r"[A-Z]\d[A-Z]\s?\d[A-Z]\d",
                    str(v_local).strip().upper(),
                ):
                    merged[k] = v_local
                else:
                    merged[k] = v_gem
                divergences += 1
                divergence_warnings.append(
                    f"Divergence sur {k} : parser local = "
                    f"« {v_local} », Gemini = « {v_gem} ». Pris "
                    f"{'parser local' if merged[k] == v_local else 'Gemini'} "
                    "par défaut. Vérifier manuellement."
                )
            else:
                # Divergence majeure générique → on prend Gemini
                # (plus robuste sur formats inhabituels) ET on
                # flag la divergence pour vérification humaine.
                merged[k] = v_gem
                divergences += 1
                divergence_warnings.append(
                    f"Divergence sur {k} : parser local = "
                    f"« {v_local} », Gemini = « {v_gem} ». Pris "
                    "Gemini par défaut. Vérifier manuellement."
                )
        elif has_local:
            merged[k] = v_local
        elif has_gem:
            merged[k] = v_gem
        # else : aucun n'a la valeur, on n'écrit rien.

    return merged, divergence_warnings, divergences


# Champs pour lesquels on conserve la trace par couche (utile pour le
# tooltip côté UI + la détection de divergence post-merge). Sous-set
# numérique principalement — les strings (description, courtier…) ne
# sont pas utiles à comparer.
_PER_SOURCE_TRACKED: Tuple[str, ...] = (
    "address", "city", "postal_code",
    "asking_price", "nb_logements", "revenus_bruts",
    "taxes_municipales", "taxes_scolaires", "assurances",
    "energie", "depenses_autres",
    "annee_construction",
    "superficie_terrain", "superficie_batiment",
    "evaluation_municipale",
    "nb_stationnements",
)


def _build_per_source(
    local: Optional[Dict[str, Any]],
    gemini: Optional[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Construit le dict ``{field: {"local": x, "gemini": y}}`` qui
    accompagnera la fiche extraite vers le validator (Phase A3).
    Garde uniquement les champs ``_PER_SOURCE_TRACKED`` où au moins
    une des deux couches a une valeur."""
    out: Dict[str, Dict[str, Any]] = {}
    loc = local or {}
    gem = gemini or {}
    for f in _PER_SOURCE_TRACKED:
        v_loc = loc.get(f)
        v_gem = gem.get(f)
        has_loc = v_loc not in (None, "", {}, [])
        has_gem = v_gem not in (None, "", {}, [])
        if has_loc or has_gem:
            entry: Dict[str, Any] = {}
            if has_loc:
                entry["local"] = v_loc
            if has_gem:
                entry["gemini"] = v_gem
            out[f] = entry
    return out


def _count_fields(d: Dict[str, Any]) -> int:
    """Nb de champs non-vides dans un dict (exclut source_url qui
    est de la métadonnée, pas un champ extrait)."""
    return sum(
        1
        for k, v in d.items()
        if k != "source_url" and v not in (None, "", {}, [])
    )


def _select_model_used(
    n_local: int, n_gemini: int, gemini_skipped: bool
) -> str:
    """Sélectionne la valeur de ``model_used`` selon ce que chaque
    couche a produit.

    - ``"local + gemini"`` : les deux ont sorti au moins un champ.
    - ``"local"``          : seul le parser local a sorti des champs
      (Gemini désactivé, en erreur, ou rien renvoyé).
    - ``"gemini"``         : seul Gemini a sorti des champs.
    - ``"none"``           : aucune des deux couches n'a rien.
    """
    if n_local > 0 and n_gemini > 0:
        return "local + gemini"
    if n_local > 0:
        return "local"
    if n_gemini > 0:
        return "gemini"
    return "none"


# ── Pipeline principal ────────────────────────────────────────────


async def _run_gemini_safely(
    material: str,
    images: List[Tuple[str, bytes]],
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
    """Wrap ``_gemini_extract`` pour ne JAMAIS lever — toute exception
    est convertie en ``(None, raison)`` pour ne pas faire échouer
    l'extraction si Gemini tombe (quota, réseau, JSON invalide…).

    Le caller émet alors un warning visible à l'utilisateur et on
    poursuit avec le résultat du parser local seul."""
    try:
        return await _gemini_extract(material, images)
    except Exception as exc:  # noqa: BLE001
        log.warning(
            "Gemini extraction a levé inattenduement : %s (%s)",
            type(exc).__name__,
            exc,
        )
        return None, f"Gemini injoignable ({type(exc).__name__})"


async def extract_lead_info(
    *,
    urls: List[str] | None = None,
    text: str | None = None,
    files: List[Tuple[str, str, bytes]] | None = None,
) -> ExtractionResult:
    """Pipeline principal d'extraction — parser local + Gemini EN
    PARALLÈLE avec merge intelligent par champ.

    Stratégie (Phase A1, voir docstring module pour le détail) :
      1. Préparation : fetch URLs, OCR PDF/images, lecture Excel.
         C'est ici que les I/O lourds sont concentrés. Les sources
         locales (parser dédié, JSON-LD, regex texte) sont accumulées
         dans ``extracted`` ; le matériel pour Gemini est préparé en
         parallèle (HTML strippé, textes OCR, textes Excel,
         images natives transmises sans OCR).
      2. ``asyncio.gather`` lance EN PARALLÈLE :
         (a) la finalisation du parser local (synchrone — regroupement
             par adresse, merge multi-sources) wrappée dans
             ``asyncio.to_thread`` pour ne pas bloquer la boucle.
         (b) ``_gemini_extract`` sur tout le matériel consolidé.
      3. Merge intelligent local↔Gemini par champ (cf.
         ``_merge_local_gemini``) :
           - concordance → confiance forte, on garde local.
           - une seule source a la valeur → on prend cette source.
           - divergence majeure → Gemini par défaut + warning, sauf
             adresse locale clairement valide → on garde local.
      4. ``model_used`` reflète la cascade réelle (``"local + gemini"``
         / ``"local"`` / ``"gemini"`` / ``"none"``).
    """
    warnings: List[str] = []
    # Liste de tuples (parser_tag, addr_key, data_dict).
    extracted: List[Tuple[str, str, Dict[str, Any]]] = []
    # Matériel brut consolidé pour l'extraction IA en parallèle
    # (Gemini reçoit TOUTE la matière, pas juste un fallback).
    gemini_material_parts: List[str] = []
    gemini_images: List[Tuple[str, bytes]] = []

    # ── URLs ───
    for u in urls or []:
        u = (u or "").strip()
        if not u:
            continue
        domain = urlparse(u).netloc
        parser_kind = _parser_for_domain(domain)
        html, fetch_err = await _fetch_html_rendered(u, parser_kind)
        if fetch_err:
            warnings.append(f"URL {u} : {fetch_err}")
            continue
        if parser_kind == "centris" and not scraping_proxy.vps_available():
            warnings.append(
                f"URL {u} : la section « Détails financiers » de "
                "Centris (évaluation municipale, taxes, dépenses) est "
                "chargée en JavaScript et reste invisible sans rendu "
                "navigateur. Active le VPS de scraping "
                "(SCRAPING_VPS_URL / SCRAPING_VPS_KEY) pour l'importer."
            )
        gemini_material_parts.append(
            f"[Page web : {u}]\n{_strip_html(html)[:40_000]}"
        )

        sources_for_url: List[Tuple[str, Dict[str, Any]]] = []
        try:
            if parser_kind == "centris":
                sources_for_url.append(("centris", parse_centris(html)))
                # Le DOM Centris change souvent → on complète (et on
                # priorise) avec l'analyse texte de la page strippée,
                # ancrée sur les libellés stables de Centris.
                sources_for_url.append(
                    ("centris-text", parse_text(_strip_html(html)))
                )
            elif parser_kind == "duproprio":
                sources_for_url.append(("duproprio", parse_duproprio(html)))
            elif parser_kind == "realtor":
                sources_for_url.append(("realtor", parse_realtor(html)))
            else:
                # Site générique : on essaie __NEXT_DATA__ d'abord
                # (couvre pmml.ca et tout autre Next.js), puis JSON-LD,
                # puis regex sur HTML stripped.
                if has_next_data(html):
                    sources_for_url.append(
                        ("next_data", parse_next_data(html))
                    )
                ld = parse_jsonld(html)
                if ld:
                    sources_for_url.append(("jsonld", ld))
                if not sources_for_url:
                    # Ni Next ni JSON-LD : full regex.
                    sources_for_url.append(
                        ("text", parse_text(_strip_html(html)))
                    )
                else:
                    # Toujours compléter avec le texte strippé pour
                    # rattraper taxes, typologie québécoise, etc.
                    sources_for_url.append(
                        ("text", parse_text(_strip_html(html)))
                    )
        except Exception as exc:  # noqa: BLE001
            log.exception("parser failed for %s", u)
            warnings.append(
                f"URL {u} : erreur de parsing ({type(exc).__name__})"
            )
            continue

        # Mini-merge intra-URL.
        url_merged = merge_extracted(sources_for_url)
        if not url_merged:
            warnings.append(
                f"URL {u} : aucun champ extrait — page sans données "
                "structurées reconnaissables (Cloudflare, SPA non "
                "Next.js, ou format inhabituel)"
            )
            continue
        url_merged["source_url"] = u
        addr_key = _addr_key(url_merged)
        extracted.append((f"url:{parser_kind}", addr_key, url_merged))

    # ── Texte libre ───
    if text and text.strip():
        gemini_material_parts.append(f"[Texte fourni]\n{text.strip()}")
        td = parse_text(text)
        if td:
            extracted.append(("text", _addr_key(td), td))
        else:
            warnings.append(
                "Texte libre : aucun champ reconnaissable extrait"
            )

    # ── Fichiers ───
    for filename, content_type, blob in files or []:
        ct = (content_type or "").lower()
        if ct == "application/pdf" or filename.lower().endswith(".pdf"):
            # 1) Couche texte native via pypdf (PDFs descriptifs MLS,
            #    courtiers qui exportent depuis Centris, etc.).
            pdf_text = parse_pdf(blob)
            ocr_used = False
            if (
                not pdf_text.strip()
                or len(pdf_text.strip()) < _PDF_OCR_FALLBACK_THRESHOLD
            ):
                # 2) PDF probablement scanné (pas de couche texte ou
                #    presque rien). Fallback OCR Tesseract page par page.
                log.info(
                    "PDF '%s' : couche texte vide/courte (%d chars), "
                    "fallback OCR Tesseract",
                    filename,
                    len(pdf_text.strip()),
                )
                ocr_text = parse_pdf_ocr(blob, filename=filename)
                if ocr_text.strip():
                    pdf_text = _normalize_ocr_text(ocr_text)
                    ocr_used = True
            if not pdf_text.strip():
                warnings.append(
                    f"PDF « {filename} » : ni texte natif ni OCR "
                    "exploitable (PDF illisible ou binaires Tesseract/"
                    "poppler indisponibles sur le serveur)"
                )
                continue
            gemini_material_parts.append(
                f"[PDF : {filename}]\n{pdf_text[:40_000]}"
            )
            td = parse_text(pdf_text)
            if td:
                tag = "pdf-ocr" if ocr_used else "pdf"
                extracted.append((tag, _addr_key(td), td))
            else:
                preview_pdf = (pdf_text or "").strip().replace("\n", " ⏎ ")[:300]
                warnings.append(
                    f"PDF « {filename} » : texte extrait "
                    f"({'OCR' if ocr_used else 'natif'}, "
                    f"{len(pdf_text or '')} chars) mais aucun champ "
                    f"reconnaissable. Aperçu : « {preview_pdf}… »"
                )
        elif ct.startswith("image/") or filename.lower().endswith(
            (".png", ".jpg", ".jpeg", ".heic", ".heif", ".webp", ".tiff", ".bmp")
        ):
            # Image transmise telle quelle à Gemini (lecture native,
            # bien meilleure que l'OCR). L'OCR reste calculé comme
            # filet de secours du parser local.
            gemini_images.append((ct or "image/png", blob))
            # Screenshot de tableau Excel, photo de fiche MLS, capture
            # de courriel, photo HEIC iPhone, etc. → OCR Tesseract.
            ocr_text = parse_image_ocr(blob, filename=filename)
            if not ocr_text.strip():
                tess_status = _check_tesseract_status()
                warnings.append(
                    f"Image « {filename} » : OCR n'a rien extrait. "
                    f"État Tesseract serveur : {tess_status}"
                )
                continue
            normalized = _normalize_ocr_text(ocr_text)
            td = parse_text(normalized)
            preview = ocr_text.strip().replace("\n", " ⏎ ")[:300]
            if td:
                extracted.append(("image-ocr", _addr_key(td), td))
                warnings.append(
                    f"Image « {filename} » : OCR OK "
                    f"({len(ocr_text)} chars extraits, "
                    f"{len(td)} champs reconnus). Aperçu : « {preview}… »"
                )
            else:
                warnings.append(
                    f"Image « {filename} » : OCR a extrait "
                    f"{len(ocr_text)} chars mais aucun champ "
                    f"reconnaissable. Texte OCR (aperçu) : « {preview}… »"
                )
        elif (
            "excel" in ct
            or "spreadsheetml" in ct
            or filename.lower().endswith((".xlsx", ".xls"))
        ):
            # Excel — Phil reçoit des listes d'immeubles à vendre
            # exportées par des courtiers en .xlsx. On convertit le
            # tableau en texte structuré (headers + lignes) puis on
            # passe à parse_text ET à Gemini en parallèle.
            xl_text = parse_excel(blob, filename=filename)
            if not xl_text.strip():
                warnings.append(
                    f"Excel « {filename} » : impossible d'extraire "
                    "le contenu (fichier illisible ou .xls sans "
                    "xlrd installé — convertir en .xlsx pour le "
                    "moment)."
                )
                continue
            gemini_material_parts.append(
                f"[Excel : {filename}]\n{xl_text[:60_000]}"
            )
            td = parse_text(xl_text)
            if td:
                extracted.append(("excel", _addr_key(td), td))
            else:
                preview_xl = xl_text.strip().replace("\n", " ⏎ ")[:300]
                warnings.append(
                    f"Excel « {filename} » : tableau lu "
                    f"({len(xl_text)} chars) mais aucun champ "
                    f"reconnaissable par le parser local. "
                    f"Aperçu : « {preview_xl}… » — Gemini est tout "
                    "de même appelé en parallèle sur ce contenu."
                )
        else:
            warnings.append(
                f"Fichier « {filename} » : type {ct or 'inconnu'} "
                "non supporté"
            )

    # ── Lancement EN PARALLÈLE du finalize local + Gemini ───
    # Le parser local est synchrone (CPU) — on le wrappe dans
    # asyncio.to_thread pour ne pas bloquer la boucle pendant que
    # Gemini fait son appel réseau. Les deux tâches s'exécutent
    # simultanément ; on les attend via asyncio.gather.

    def _finalize_local() -> List[Dict[str, Any]]:
        """Regroupement par adresse + merge multi-sources locales.
        Synchrone — appelé via asyncio.to_thread."""
        by_addr: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
        for tag, addr_key, data in extracted:
            by_addr.setdefault(addr_key, []).append(
                (tag.split(":")[-1], data)
            )
        out: List[Dict[str, Any]] = []
        for _addr_k, sources in by_addr.items():
            merged = merge_extracted(sources)
            if merged:
                out.append(merged)
        return out

    gemini_material = "\n\n".join(
        p for p in gemini_material_parts if p.strip()
    )

    # Pas de matière pour Gemini (aucun input net) → on ne l'appelle
    # pas du tout. Sinon on le lance EN PARALLÈLE du finalize local.
    if gemini_material.strip() or gemini_images:
        local_task = asyncio.to_thread(_finalize_local)
        gemini_task = _run_gemini_safely(gemini_material, gemini_images)
        local_data_list, gemini_result = await asyncio.gather(
            local_task, gemini_task
        )
        gemini_data, gemini_err = gemini_result
    else:
        local_data_list = _finalize_local()
        gemini_data, gemini_err = None, None

    # Warning Gemini indisponible — émis seulement si on l'a tenté
    # et qu'il a échoué pour une vraie raison (pas la matière vide).
    if (
        gemini_data is None
        and gemini_err
        and (gemini_material.strip() or gemini_images)
    ):
        warnings.append(
            f"Gemini indisponible ({gemini_err}) — extraction sur "
            "parser local seul."
        )

    # ── Merge local ↔ Gemini par adresse ───
    gemini_list = gemini_data or []
    # Indexe Gemini par clé d'adresse pour fusionner avec le bon
    # immeuble local quand il y en a plusieurs.
    gemini_by_addr: Dict[str, Dict[str, Any]] = {}
    for g in gemini_list:
        if not isinstance(g, dict):
            continue
        gk = _addr_key(g)
        # Si la même adresse apparaît plusieurs fois côté Gemini, on
        # union (gemini répond rarement avec doublons mais sait-on
        # jamais).
        if gk in gemini_by_addr:
            gemini_by_addr[gk] = {**gemini_by_addr[gk], **g}
        else:
            gemini_by_addr[gk] = g

    data_out: List[Dict[str, Any]] = []
    # Phase A3 : pour chaque fiche de ``data_out``, on conserve la
    # valeur vue par chaque couche (avant merge) — sert au
    # `lead_validation.validate_extraction()` pour détecter les
    # divergences et alimenter le tooltip côté UI.
    per_source_values_out: List[Dict[str, Dict[str, Any]]] = []
    total_divergences = 0
    used_gemini_keys: set = set()

    for loc in local_data_list:
        loc_key = _addr_key(loc)
        gem = gemini_by_addr.get(loc_key)
        if gem is None and len(gemini_by_addr) == 1 and loc_key == "unknown":
            # Cas standard : 1 seul immeuble, pas d'adresse côté
            # parser local → on associe l'unique fiche Gemini.
            (only_key,) = gemini_by_addr.keys()
            gem = gemini_by_addr[only_key]
            used_gemini_keys.add(only_key)
        elif gem is None and len(gemini_by_addr) == 1 and len(local_data_list) == 1:
            # 1 ↔ 1 sans correspondance d'adresse exacte → on fait
            # tout de même le merge (Gemini a pu trouver une
            # adresse que le local a ratée, ou inversement).
            (only_key,) = gemini_by_addr.keys()
            gem = gemini_by_addr[only_key]
            used_gemini_keys.add(only_key)
        else:
            used_gemini_keys.add(loc_key)

        if gem:
            merged, div_warns, n_div = _merge_local_gemini(loc, gem)
            warnings.extend(div_warns)
            total_divergences += n_div
            data_out.append(merged)
            per_source_values_out.append(_build_per_source(loc, gem))
        else:
            data_out.append(loc)
            per_source_values_out.append(_build_per_source(loc, None))

    # Fiches Gemini sans pendant local (Gemini a détecté un immeuble
    # que le parser local n'a pas su extraire) → on les ajoute.
    for gk, gem in gemini_by_addr.items():
        if gk in used_gemini_keys:
            continue
        data_out.append(dict(gem))
        per_source_values_out.append(_build_per_source(None, gem))

    # ── Logging détaillé ───
    n_local_total = sum(_count_fields(d) for d in local_data_list)
    n_gemini_total = sum(
        _count_fields(g) for g in gemini_list if isinstance(g, dict)
    )
    n_merged_total = sum(_count_fields(d) for d in data_out)
    sources_tags = []
    if urls:
        sources_tags.append("url")
    if text and text.strip():
        sources_tags.append("texte")
    if files:
        sources_tags.append("fichier")
    src_label = "+".join(sources_tags) or "vide"
    log.info(
        "Extraction %s : local=%d champs, gemini=%d champs, "
        "merged=%d champs, %d divergence(s)",
        src_label,
        n_local_total,
        n_gemini_total,
        n_merged_total,
        total_divergences,
    )

    # ── Sélection model_used et warnings de couverture ───
    if data_out:
        best = max((_coverage(r) for r in data_out), default=0)
        if best < 4:
            warnings.append(
                f"⚠ Extraction faible — {best} champ(s) clé(s) sur "
                f"{len(_KEY_FIELDS)} reconnus. Ce format de "
                "document est mal pris en charge même avec Gemini : "
                "transmets-le à l'équipe pour le faire évoluer. En "
                "attendant, complète les champs manquants à la main."
            )
        model_used = _select_model_used(
            n_local_total, n_gemini_total, gemini_skipped=False
        )
        return ExtractionResult(
            data=data_out,
            model_used=model_used,
            warnings=warnings,
            per_source_values=per_source_values_out,
        )

    if not warnings:
        warnings.append("Aucune donnée exploitable extraite.")
    return ExtractionResult(
        data=[],
        model_used=_select_model_used(0, 0, gemini_skipped=False),
        warnings=warnings,
    )


def _addr_key(d: Dict[str, Any]) -> str:
    """Clé de regroupement par adresse. Si pas d'adresse, on retombe
    sur un singleton « unknown » — toutes les sources sans adresse
    seront fusionnées en une seule fiche (cas standard pour 99% des
    inputs)."""
    addr = (d.get("address") or "").strip().lower()
    if not addr:
        return "unknown"
    # Normalise : retire les espaces multiples, garde alphanumeric.
    return re.sub(r"[^\w]+", "_", addr)[:80]
