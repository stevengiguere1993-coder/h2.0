"""Import Excel d'un immeuble complet — modèle + parsing.

Phil (2026-07-14) : « un formulaire téléchargeable, que je peux remplir
avec toutes les informations d'un immeuble, ses logements, locataires,
hypothèques… et qu'on peut importer pour créer l'immeuble au complet ».

Le modèle .xlsx contient 5 feuilles :
    Instructions        — mode d'emploi
    Immeuble            — fiche verticale (une valeur par ligne)
    Hypothèques         — un prêt par ligne
    Logements           — un logement par ligne
    Locataires & baux   — un locataire + son bail par ligne (rattaché au
                          logement par son numéro)

Le parsing est TOUT-OU-RIEN : chaque problème est collecté avec sa
feuille/ligne ; s'il y a la moindre erreur on ne crée rien et on renvoie
la liste complète (l'utilisateur corrige son fichier et réessaie).
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constantes du modèle ────────────────────────────────────────────────

SHEET_INSTRUCTIONS = "Instructions"
SHEET_IMMEUBLE = "Immeuble"
SHEET_HYPOTHEQUES = "Hypothèques"
SHEET_LOGEMENTS = "Logements"
SHEET_BAUX = "Locataires & baux"

# Feuille Immeuble : (libellé affiché, clé interne, requis, aide)
IMMEUBLE_ROWS: list[tuple[str, str, bool, str]] = [
    ("Adresse", "address", True, "ex. 1234 rue Ontario Est"),
    ("Ville", "city", False, "ex. Montréal"),
    ("Code postal", "postal_code", False, "ex. H2L 1S6"),
    (
        "Type",
        "type",
        False,
        "residentiel / commercial / mixte / unifamilial / autre "
        "(défaut : residentiel)",
    ),
    ("Année de construction", "annee_construction", False, "ex. 1985"),
    ("Nombre de logements (déclaré)", "nb_logements", False, "ex. 6"),
    ("Prix d'achat ($)", "purchase_price", False, "ex. 850000"),
    ("Date d'achat", "purchase_date", False, "AAAA-MM-JJ"),
    (
        "Gestion externe",
        "gestion_externe",
        False,
        "OUI si géré par une compagnie externe, sinon vide/NON",
    ),
    ("Nom du gestionnaire externe", "gestionnaire_externe_nom", False, ""),
    (
        "Contact du gestionnaire externe",
        "gestionnaire_externe_contact",
        False,
        "courriel ou téléphone",
    ),
]

HYPO_HEADERS: list[tuple[str, str]] = [
    ("Rang", "1 = première hypothèque (défaut 1)"),
    ("Prêteur *", "ex. Desjardins"),
    ("Montant initial ($) *", "ex. 600000"),
    ("Balance actuelle ($)", "vide = calcul automatique"),
    ("Taux (%)", "ex. 4.89"),
    ("Type de taux", "fixe / variable (défaut fixe)"),
    (
        "Composition des intérêts",
        "semi-annuelle (résidentiel, défaut) / mensuelle (commercial)",
    ),
    ("Amortissement (années)", "ex. 25"),
    ("Date de début", "AAAA-MM-JJ"),
    ("Date de fin du terme", "AAAA-MM-JJ"),
    ("Paiement mensuel ($)", "vide = calcul automatique"),
]

LOGEMENT_HEADERS: list[tuple[str, str]] = [
    ("Numéro *", "ex. 101, A, 2e étage gauche"),
    ("Pièces", "ex. 3.5 pour un 3½"),
    ("Chambres", "ex. 2"),
    ("Salles de bain", "ex. 1 ou 1.5"),
    ("Superficie (pi²)", "ex. 750"),
    ("Étage", "ex. 1"),
    ("Loyer demandé ($)", "référence pour les annonces"),
    ("Notes", ""),
]

BAIL_HEADERS: list[tuple[str, str]] = [
    ("Logement (numéro) *", "doit exister dans la feuille Logements"),
    ("Nom complet du locataire *", "ex. Jean Tremblay"),
    ("Courriel", ""),
    ("Téléphone", ""),
    ("Loyer mensuel ($) *", "ex. 1250"),
    ("Début du bail *", "AAAA-MM-JJ"),
    ("Fin du bail *", "AAAA-MM-JJ"),
    ("Dépôt de garantie ($)", "vide si aucun"),
    ("Notes", ""),
]

_TYPES_IMMEUBLE = {"residentiel", "commercial", "mixte", "unifamilial", "autre"}

_HDR_FILL = PatternFill("solid", fgColor="1F2937")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HELP_FONT = Font(italic=True, color="6B7280", size=9)
_REQ_FONT = Font(bold=True)


def _norm(s: Any) -> str:
    """Normalise un libellé : minuscules, sans accents ni ponctuation."""
    txt = str(s or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", txt).strip()


# ── Génération du modèle ────────────────────────────────────────────────


def generate_template() -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = SHEET_INSTRUCTIONS
    lignes = [
        "IMPORT D'UN IMMEUBLE COMPLET — MODE D'EMPLOI",
        "",
        "1. Remplis la feuille « Immeuble » (colonne B). Seule l'adresse est obligatoire.",
        "2. Feuille « Hypothèques » : un prêt par ligne (facultatif).",
        "3. Feuille « Logements » : un logement par ligne (numéro obligatoire).",
        "4. Feuille « Locataires & baux » : un locataire + son bail par ligne.",
        "   La colonne « Logement (numéro) » doit correspondre à un numéro",
        "   de la feuille Logements.",
        "5. Les dates s'écrivent AAAA-MM-JJ (ou en format date d'Excel).",
        "6. Les colonnes marquées * sont obligatoires ; les lignes d'exemple",
        "   en gris peuvent être écrasées ou supprimées.",
        "7. Retourne dans Kratos → Gestion immobilière → Immeubles →",
        "   « Importer (Excel) », choisis l'entreprise propriétaire et téléverse",
        "   le fichier. Rien n'est créé s'il reste des erreurs (elles te seront",
        "   listées avec la feuille et la ligne).",
    ]
    for i, txt in enumerate(lignes, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        if i == 1:
            c.font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 90

    ws = wb.create_sheet(SHEET_IMMEUBLE)
    ws.cell(row=1, column=1, value="Champ").font = _HDR_FONT
    ws.cell(row=1, column=2, value="Valeur").font = _HDR_FONT
    ws.cell(row=1, column=3, value="Aide").font = _HDR_FONT
    for col in (1, 2, 3):
        ws.cell(row=1, column=col).fill = _HDR_FILL
    for i, (label, _key, requis, aide) in enumerate(IMMEUBLE_ROWS, start=2):
        c = ws.cell(row=i, column=1, value=f"{label} *" if requis else label)
        if requis:
            c.font = _REQ_FONT
        h = ws.cell(row=i, column=3, value=aide)
        h.font = _HELP_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55

    def _sheet_tabulaire(titre: str, headers: list[tuple[str, str]]) -> None:
        # Pas de ligne d'exemple avec des valeurs réalistes : elle serait
        # importée telle quelle si l'utilisateur oublie de l'effacer. La
        # ligne d'aide (détectée et sautée au parsing) donne les « ex. ».
        w = wb.create_sheet(titre)
        for j, (head, aide) in enumerate(headers, start=1):
            c = w.cell(row=1, column=j, value=head)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL
            c.alignment = Alignment(wrap_text=True, vertical="top")
            a = w.cell(row=2, column=j, value=aide)
            a.font = _HELP_FONT
            w.column_dimensions[get_column_letter(j)].width = max(
                16, min(34, len(head) + 4)
            )
        w.freeze_panes = "A3"

    _sheet_tabulaire(SHEET_HYPOTHEQUES, HYPO_HEADERS)
    _sheet_tabulaire(SHEET_LOGEMENTS, LOGEMENT_HEADERS)
    _sheet_tabulaire(SHEET_BAUX, BAIL_HEADERS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Parsing ─────────────────────────────────────────────────────────────


@dataclass
class HypoImport:
    rang: int
    preteur: str
    montant_initial: float
    balance_actuelle: Optional[float]
    taux_pct: Optional[float]
    type_taux: Optional[str]
    composition_interets: str
    amortissement_mois: Optional[int]
    date_debut: Optional[date]
    date_fin_terme: Optional[date]
    paiement_mensuel: Optional[float]


@dataclass
class LogementImport:
    numero: str
    nb_pieces_decimal: Optional[float]
    nb_chambres: Optional[int]
    nb_sdb: Optional[float]
    superficie_pi2: Optional[float]
    etage: Optional[int]
    loyer_demande: Optional[float]
    notes: Optional[str]


@dataclass
class BailImport:
    logement_numero: str
    full_name: str
    email: Optional[str]
    phone: Optional[str]
    loyer_mensuel: float
    date_debut: date
    date_fin: date
    depot_garantie: Optional[float]
    notes: Optional[str]


@dataclass
class ImmeubleImport:
    immeuble: dict[str, Any]
    hypotheques: list[HypoImport] = field(default_factory=list)
    logements: list[LogementImport] = field(default_factory=list)
    baux: list[BailImport] = field(default_factory=list)


class ImportErreurs(Exception):
    """Fichier invalide — `erreurs` liste tous les problèmes trouvés."""

    def __init__(self, erreurs: list[str]):
        self.erreurs = erreurs
        super().__init__("; ".join(erreurs))


def _as_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _as_float(v: Any, err: list[str], ctx: str) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        # Tolère « 1 250,50 $ » saisi en texte.
        if isinstance(v, str):
            v = v.replace("$", "").replace(" ", "").replace(" ", "")
            v = v.replace(",", ".")
        f = float(v)
        if f < 0:
            err.append(f"{ctx} : valeur négative.")
            return None
        return f
    except (TypeError, ValueError):
        err.append(f"{ctx} : nombre invalide ({v!r}).")
        return None


def _as_int(v: Any, err: list[str], ctx: str) -> Optional[int]:
    f = _as_float(v, err, ctx)
    return int(round(f)) if f is not None else None


def _as_date(v: Any, err: list[str], ctx: str) -> Optional[date]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    err.append(f"{ctx} : date invalide ({s!r}) — format attendu AAAA-MM-JJ.")
    return None


def _as_bool(v: Any) -> bool:
    return _norm(v) in {"oui", "yes", "true", "vrai", "x", "1"}


def _ligne_vide(cells: tuple) -> bool:
    return all(
        c is None or (isinstance(c, str) and not c.strip()) for c in cells
    )


def _est_ligne_aide(cells: tuple, headers: list[tuple[str, str]]) -> bool:
    """La ligne 2 du modèle contient les textes d'aide — on la saute si
    elle n'a pas été écrasée (première cellule = texte d'aide connu)."""
    premier = _norm(cells[0]) if cells else ""
    aides = {_norm(a) for _h, a in headers if a}
    return bool(premier) and premier in aides


def parse_workbook(blob: bytes) -> ImmeubleImport:
    """Parse le fichier rempli. Lève ImportErreurs si quoi que ce soit
    cloche — dans ce cas RIEN ne doit être créé."""
    err: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        raise ImportErreurs(
            ["Fichier illisible — utilise le modèle .xlsx fourni."]
        )

    noms = {_norm(n): n for n in wb.sheetnames}

    def _ws(titre: str):
        return wb[noms[_norm(titre)]] if _norm(titre) in noms else None

    # ── Immeuble (feuille verticale) ──
    ws = _ws(SHEET_IMMEUBLE)
    if ws is None:
        raise ImportErreurs(
            [f"Feuille « {SHEET_IMMEUBLE} » introuvable — utilise le modèle fourni."]
        )
    par_label = { _norm(label): key for label, key, _r, _a in IMMEUBLE_ROWS }
    imm: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        label = _norm(re.sub(r"\*", "", str(row[0] or "")))
        if not label or label not in par_label:
            continue
        imm[par_label[label]] = row[1]

    address = _as_str(imm.get("address"))
    if not address:
        err.append("Immeuble : l'adresse est obligatoire.")
    type_norm = _norm(imm.get("type")) or "residentiel"
    type_norm = type_norm.replace(" ", "")
    if type_norm not in _TYPES_IMMEUBLE:
        err.append(
            f"Immeuble : type invalide ({imm.get('type')!r}) — "
            "residentiel / commercial / mixte / unifamilial / autre."
        )
        type_norm = "residentiel"

    immeuble_data: dict[str, Any] = {
        "address": address or "",
        "city": _as_str(imm.get("city")),
        "postal_code": _as_str(imm.get("postal_code")),
        "type": type_norm,
        "annee_construction": _as_int(
            imm.get("annee_construction"), err, "Immeuble / année"
        ),
        "nb_logements": _as_int(
            imm.get("nb_logements"), err, "Immeuble / nombre de logements"
        ),
        "purchase_price": _as_float(
            imm.get("purchase_price"), err, "Immeuble / prix d'achat"
        ),
        "purchase_date": _as_date(
            imm.get("purchase_date"), err, "Immeuble / date d'achat"
        ),
        "gestion_externe": _as_bool(imm.get("gestion_externe")),
        "gestionnaire_externe_nom": _as_str(
            imm.get("gestionnaire_externe_nom")
        ),
        "gestionnaire_externe_contact": _as_str(
            imm.get("gestionnaire_externe_contact")
        ),
    }

    out = ImmeubleImport(immeuble=immeuble_data)

    # ── Hypothèques ──
    ws = _ws(SHEET_HYPOTHEQUES)
    if ws is not None:
        for i, row in enumerate(
            ws.iter_rows(min_row=2, max_col=len(HYPO_HEADERS), values_only=True),
            start=2,
        ):
            if _ligne_vide(row) or _est_ligne_aide(row, HYPO_HEADERS):
                continue
            ctx = f"Hypothèques ligne {i}"
            preteur = _as_str(row[1])
            montant = _as_float(row[2], err, f"{ctx} / montant initial")
            if not preteur:
                err.append(f"{ctx} : le prêteur est obligatoire.")
            if montant is None or montant <= 0:
                err.append(f"{ctx} : montant initial obligatoire (> 0).")
            compo = _norm(row[6])
            composition = (
                "mensuelle" if compo.startswith("mensuel") else "semi"
            )
            type_taux = _norm(row[5]) or None
            if type_taux and type_taux not in {"fixe", "variable"}:
                err.append(f"{ctx} : type de taux invalide ({row[5]!r}).")
                type_taux = None
            amort_annees = _as_float(
                row[7], err, f"{ctx} / amortissement (années)"
            )
            out.hypotheques.append(
                HypoImport(
                    rang=_as_int(row[0], err, f"{ctx} / rang") or 1,
                    preteur=preteur or "",
                    montant_initial=montant or 0.0,
                    balance_actuelle=_as_float(
                        row[3], err, f"{ctx} / balance"
                    ),
                    taux_pct=_as_float(row[4], err, f"{ctx} / taux"),
                    type_taux=type_taux or "fixe",
                    composition_interets=composition,
                    amortissement_mois=(
                        int(round(amort_annees * 12))
                        if amort_annees
                        else None
                    ),
                    date_debut=_as_date(
                        row[8], err, f"{ctx} / date de début"
                    ),
                    date_fin_terme=_as_date(
                        row[9], err, f"{ctx} / fin du terme"
                    ),
                    paiement_mensuel=_as_float(
                        row[10], err, f"{ctx} / paiement mensuel"
                    ),
                )
            )

    # ── Logements ──
    ws = _ws(SHEET_LOGEMENTS)
    numeros: set[str] = set()
    if ws is not None:
        for i, row in enumerate(
            ws.iter_rows(
                min_row=2, max_col=len(LOGEMENT_HEADERS), values_only=True
            ),
            start=2,
        ):
            if _ligne_vide(row) or _est_ligne_aide(row, LOGEMENT_HEADERS):
                continue
            ctx = f"Logements ligne {i}"
            numero = _as_str(row[0])
            if not numero:
                err.append(f"{ctx} : le numéro est obligatoire.")
                continue
            if numero.lower() in numeros:
                err.append(f"{ctx} : numéro en double ({numero}).")
                continue
            numeros.add(numero.lower())
            out.logements.append(
                LogementImport(
                    numero=numero,
                    nb_pieces_decimal=_as_float(
                        row[1], err, f"{ctx} / pièces"
                    ),
                    nb_chambres=_as_int(row[2], err, f"{ctx} / chambres"),
                    nb_sdb=_as_float(row[3], err, f"{ctx} / salles de bain"),
                    superficie_pi2=_as_float(
                        row[4], err, f"{ctx} / superficie"
                    ),
                    etage=_as_int(row[5], err, f"{ctx} / étage"),
                    loyer_demande=_as_float(
                        row[6], err, f"{ctx} / loyer demandé"
                    ),
                    notes=_as_str(row[7]),
                )
            )

    # ── Locataires & baux ──
    ws = _ws(SHEET_BAUX)
    if ws is not None:
        for i, row in enumerate(
            ws.iter_rows(min_row=2, max_col=len(BAIL_HEADERS), values_only=True),
            start=2,
        ):
            if _ligne_vide(row) or _est_ligne_aide(row, BAIL_HEADERS):
                continue
            ctx = f"Locataires & baux ligne {i}"
            log_num = _as_str(row[0])
            nom = _as_str(row[1])
            loyer = _as_float(row[4], err, f"{ctx} / loyer")
            debut = _as_date(row[5], err, f"{ctx} / début du bail")
            fin = _as_date(row[6], err, f"{ctx} / fin du bail")
            if not log_num:
                err.append(f"{ctx} : le numéro de logement est obligatoire.")
            elif log_num.lower() not in numeros:
                err.append(
                    f"{ctx} : logement « {log_num} » absent de la feuille "
                    "Logements."
                )
            if not nom:
                err.append(f"{ctx} : le nom du locataire est obligatoire.")
            if loyer is None or loyer <= 0:
                err.append(f"{ctx} : loyer mensuel obligatoire (> 0).")
            if debut is None:
                err.append(f"{ctx} : date de début obligatoire.")
            if fin is None:
                err.append(f"{ctx} : date de fin obligatoire.")
            if debut and fin and fin <= debut:
                err.append(f"{ctx} : la fin du bail précède son début.")
            out.baux.append(
                BailImport(
                    logement_numero=log_num or "",
                    full_name=nom or "",
                    email=_as_str(row[2]),
                    phone=_as_str(row[3]),
                    loyer_mensuel=loyer or 0.0,
                    date_debut=debut or date.today(),
                    date_fin=fin or date.today(),
                    depot_garantie=_as_float(
                        row[7], err, f"{ctx} / dépôt"
                    ),
                    notes=_as_str(row[8]),
                )
            )

    if err:
        raise ImportErreurs(err)
    return out
